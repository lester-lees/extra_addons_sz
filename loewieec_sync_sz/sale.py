# -*- encoding: utf-8 -*-
from openerp.osv import fields,osv
from openpyxl.reader.excel import load_workbook 
from openerp import tools
from openerp.tools.translate import _
import os
import re
import datetime
import logging
_logger = logging.getLogger(__name__)

class loewie_carrier(osv.osv):
    _name = "loewie.carrier"  
	
    _columns = {
        'name': fields.char(u'快递公司', required=True),
        'id_tm': fields.char(u'ID Tmall'),
        'id_jd': fields.char(u'ID JD'),	
        'reg_mail_no': fields.char(u'运单号Regex'),		
        'code_tm': fields.char(u'Code Tmall'),		
        'code_jd': fields.char(u'Code JD'),		# 京东平台，对物流公司的编码
        'base_kg':fields.float(u'基础重量',default=1),
        'base_price':fields.float(u'基础价格',default=1),
        'increment_kg':fields.float(u'增量重量',default=1),
        'increment_price': fields.float(u'增量价格',default=1),	
    }
	
class product_tmalljd_gifts(osv.osv):
    _name = "product.tmalljd.gifts"
	
    _columns = {
        'tmalljd_id': fields.many2one('product.tmalljd', 'TMallJD ID'),
        'product_id': fields.many2one('product.product','ERP Product Name'),
        'price': fields.float('Price'),		
        'qty': fields.integer('Quantity'),	
    }	
	
class product_tmalljd(osv.osv):
    _name = "product.tmalljd"

    def _get_ean13(self, cr, uid, ids, field_name, arg, context=None):	
        result = {}
        for line in self.pool.get('product.tmalljd').browse(cr, uid, ids, context=context):
            if line.erp_product_id : result[line.id] = line.erp_product_id.ean13 or line.erp_product_id.default_code
        return result		

    def _get_stock(self, cr, uid, ids, field_name, arg, context=None):	
        result = {}
        domain_products = [('location_id','=',38)]		
        quants = self.pool.get('stock.quant').read_group(cr, uid, domain_products, ['product_id', 'qty'], ['product_id'], context=context)	
        quants = dict(map(lambda x: (x['product_id'][0], x['qty']), quants))
		
        for line in self.pool.get('product.tmalljd').browse(cr, uid, ids, context=context):
            id = line.id
            if line.erp_product_id :
                pid = line.erp_product_id.id			
                result[id] = quants.get(pid, 0.0)
            else:
                result[id] = 0			
        return result		
		
    _columns = {
        'gift_ids': fields.one2many('product.tmalljd.gifts','tmalljd_id','Gift Products'), 
        'erp_product_set': fields.many2many('product.product', 'product_product_tmalljd_rel', 'tmalljd_id', 'product_id', u'Product Set(套装)'),
        'erp_product_id': fields.many2one('product.product','ERP Product Name'),
        'erp_ean13': fields.char('ERP_EAN13'), #fields.function(_get_ean13,type='char',string='ERP_EAN13'),
        'erp_stock': fields.float('ERPStock'),#fields.function(_get_stock,type='float',string='ERP库存'),		
        'ec_shop_id': fields.many2one('loewieec.shop', u'店铺'),		
        'ec_num_iid': fields.char(u'淘宝数字编码'),
        'ec_sku_id': fields.char(u'淘宝SKU_ID'),
        'ec_title':fields.char(u'商品标题'),
        'ec_price':fields.float(u'售价'),
        'ec_color':fields.char(u'颜色'),
        'ec_ean13': fields.char(u'条形码'),
        'ec_brand': fields.char(u'品牌'),
        'ec_qty': fields.integer(u'EC数量'),
        'ec_outer_code': fields.char(u'商家外部编码'),		
        'ec_product_name': fields.char(u'产品名称'),	
        'ec_product_id': fields.char(u'EC产品ID'),		
        'ec_num_custom':fields.char(u'海关代码'),	
    }
	
class loewieec_error(osv.osv):
    _name = "loewieec.error"

    _columns = {	
        'shop_id': fields.many2one('loewieec.shop', u'店铺'),
        'name': fields.char(u'错误信息'),
    }
	
class sale_coe(osv.osv):
    _name = "sale.coe"	
    _order = 'sequence asc'
	
    _columns = {
        'sequence': fields.integer(string='Sequence', default=0, readonly=True),
        'is_shuadan': fields.boolean(u'Is刷单?', default=False, readonly=True),		
        'logistic_sent': fields.boolean(u'已上传运单?', default=False),	 # , readonly=True
        'sale_id': fields.many2one('sale.order',string='销售单', readonly=True),	
        'picking_id': fields.many2one('stock.picking',string=u'出货单', readonly=True),		
        'customer': fields.many2one('res.partner',string=u'客户', readonly=True),		
        'tmi_jdi_no': fields.char(string=u'电商单号', readonly=True, states={'editable': [('readonly', False)]}),		
        'express_no': fields.char(string=u'货运单号', readonly=True, states={'editable': [('readonly', False)],'draft': [('readonly', False)]}),	
        'expresser': fields.many2one('loewie.carrier',string=u'货运公司', readonly=True, states={'editable': [('readonly', False)]}, default=3),	
        'pay_way': fields.selection([
            ('we_pay', u'包邮'),
            ('cash_pay', u'现付'),
            ('customer_pay', u'到付'),
            ], string=u'结算方式', readonly=True, states={'editable': [('readonly', False)]}),		
        'name': fields.char(string=u'收货人',required=True, readonly=True, states={'editable': [('readonly', False)]}),	
        'receive_name': fields.char(string=u'收货人old', readonly=True, states={'editable': [('readonly', False)]}),		
        'tel': fields.char(string=u'电话', readonly=True, states={'editable': [('readonly', False)]}),
        'mobile': fields.char(string=u'手机',required=True, readonly=True, states={'editable': [('readonly', False)]}),		
        'province': fields.char(string=u'省份', readonly=True, states={'editable': [('readonly', False)]}),	
        'city': fields.char(string=u'城市', readonly=True, states={'editable': [('readonly', False)]}),	
        'address': fields.char(string=u'地址',required=True, readonly=True, states={'editable': [('readonly', False)]}),	
        'zip': fields.char(string=u'邮编', readonly=True, states={'editable': [('readonly', False)]}),	
        'weight': fields.float(string=u'重量',default=1, readonly=True, states={'editable': [('readonly', False)]}),		
        'price': fields.float(string=u'快递费',default=6, readonly=True, states={'editable': [('readonly', False)]}),
        'state': fields.selection([('draft',u'草稿'),('editable',u'可编辑'),('done',u'完成')],string='State',default='draft'),		
    }
	
    def on_weight_express_change(self, cr, uid, ids, expresser=None, weight=1):
        return	
        if expresser is None: return		
        carrier = self.pool.get('loewie.carrier').browse(cr, uid, [expresser], context={})
        if weight <= carrier.base_kg:
            amount = carrier.base_price
        else:			
            round_weight = int(weight - carrier.base_kg)	
            increment_weight = 	round_weight < (weight - carrier.base_kg) and 1 or 0		
            amount = carrier.base_price + (round_weight + increment_weight) * carrier.increment_price					

        return {'value': {'price': amount}}
		
class ecommerce_orders(osv.osv):		
    _name = "ecommerce.orders"	
    _order = 'id desc'
	
    def _get_order_lines(self, cr, uid, ids, field_name, arg, context=None):	
        result = {}
        for order in self.pool.get('ecommerce.orders').browse(cr, uid, ids, context=context):
            result[order.id] = 0
			
        return result	
		
    _columns = {		
        'name': fields.char(string=u'电商单号'),
        'shop_id':fields.many2one('loewieec.shop',u'店铺名称'),	
        'partner_id':fields.many2one('res.partner',u'Partner'),			
        'logistic_sent': fields.boolean(u'已上传运单', default=False),		
        'is_shipped': fields.boolean(u'是否发货', default=False),
        'is_shuadan': fields.boolean(u'刷单', default=False),	
        'store_code': fields.char(string=u'仓库编码'),		
        'receive_name': fields.char(string=u'收件人'),	
        'pay_time': fields.datetime(u'EC支付时间'),	
        'payment': fields.float(u'买家实付'),	
        'received_payment': fields.float(u'卖家实收'),
        'consign_time': fields.datetime(u'卖家发货时间'),		
        'state': fields.char(string='Status'),
        'refund_status': fields.char(string=u'退款状态'),		
        'order_lines': fields.one2many('sale.order.line','ec_order_id',string='Order Lines'),
        'last_sync_time': fields.datetime(u'最后一次同步时间'),		
        #'order_lines': fields.function(_get_order_lines,type='one2many',relation='sale.order.line',string='Order Lines'),  refund_status
    }

	
class sale_order_line(osv.osv):
    _inherit = "sale.order.line"
	
    _columns = {
        'logistic_sent': fields.related('express_id', 'logistic_sent', type='boolean', string='已同步运单?',readonly=True),	
        'express_id': fields.many2one('sale.coe',string=u'快递信息'),	
        'tmi_jdi_no': fields.char(string=u'电商单号'),
        'ec_order_id': fields.many2one('ecommerce.orders',string=u'电商单'),		
        'buyer_nick': fields.char(u'买家昵称'),			
        'pay_time': fields.datetime(u'EC支付时间'),	
        'create_time_tmjd': fields.datetime(u'EC创建时间'),		
    }	

    def copy_sale_order_line(self, cr, uid, ids, context=None): 
        for line in self.pool.get('sale.order.line').browse(cr, uid, ids, context=context):
            line.copy()		
	
class stock_move(osv.osv):
    _inherit = "stock.move"
	
    def _get_express_id(self, cr, uid, ids, field_name, arg, context=None):	
        result = {}
        for move in self.pool.get('stock.move').browse(cr, uid, ids, context=context):
            result[move.id] = move.procurement_id.sale_line_id.express_id.id
        return result	
		
    _columns = {
        #'sale_order_line': fields.function(_get_sale_order_line, type='char',string='Sales Line'),	
        'express_id': fields.function(_get_express_id,type='many2one',relation='sale.coe',string='Express', help=u'快递信息'),
        'express_id2': fields.many2one('sale.coe',string='Express2', help=u'快递信息2'),		
    }	

class stock_picking(osv.osv):
    _inherit = "stock.picking"

    def create_return_lines_from_tmijdi_no(self, cr, uid, ids, context=None):	
		
        picking = self.browse(cr, uid, ids[0], context=context)	
        note = picking.note or ''		
        tmijdi_nos = []	
        if not note : return
        tmijdi_nos = note.strip().split(',')
        tmijdi_no_list = []		
        for tmijdi_no in tmijdi_nos:
            tmijdi_no = tmijdi_no.strip()		
            if tmijdi_no != '' : tmijdi_no_list.append( tmijdi_no )
			
        statement = "select sol.product_id, sol.product_uom_qty, so.name, sol.express_id,sol.id, sol.tmi_jdi_no from sale_order_line sol left join sale_order so on (sol.order_id=so.id) where sol.state='done' and sol.product_uom_qty>0 and sol.tmi_jdi_no in (%s)" %  ("'" + """','""".join(tmijdi_no_list)	+ "'")		
        cr.execute(statement)
        res = cr.fetchall()		
		
        vals_move = {	
            'create_uid':uid, 		
            'product_id': 0, #, 
            'product_uom_qty':0, 
            'location_dest_id':39, 
            'location_id':9, 
            'company_id':1, 
            'date':datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            'date_expected':(datetime.datetime.now() + datetime.timedelta(3)).strftime("%Y-%m-%d %H:%M:%S"), 
            'invoice_state':'none', 
            'name':'-', 
            'procure_method':'make_to_stock', 
            'state':'draft',			
            'product_uom':1, 
            'weight_uom_id':1,
            'picking_id': ids[0],			
        } 
        move_obj = self.pool.get('stock.move')
        sale_coe_obj = self.pool.get('sale.coe')		
        express_ids = []
        for item in res:
            if item[3] not in express_ids : express_ids.append(item[3])		
        express_list = {}	
        note_lines = chr(10)		
        for express_obj in sale_coe_obj.browse(cr,uid,express_ids,context=context):
            express_list[express_obj.id] = express_obj.name
            note_lines += express_obj.tmi_jdi_no + " : " + express_obj.name + ", " + 	express_obj.mobile + ", " + express_obj.express_no + ", " + express_obj.address + chr(10)	
			
        for line in res:
            val = vals_move.copy()		
            line_name = line[2] + ", "+ line[5] + ", " + express_list[line[3]]			
            val.update({'product_id':line[0],'product_uom_qty':line[1], 'name':line_name, 'sale_line_loewie':line[4]})	
            move_obj.create(cr,uid,val,context=context)
	
        picking.write({'note':note + note_lines})
        return True		
	
    def get_full_path(self, cr, uid, path):
        # sanitize ath
        path = re.sub('[.]', '', path)
        path = path.strip('/\\')
        return os.path.join(tools.config.filestore(cr.dbname), path)
		
    def import_moves_from_excel(self, cr, uid, ids, context=None):
        attachment_obj = self.pool.get('ir.attachment')
        attachment_id = attachment_obj.search(cr,uid,[('res_id', '=', ids[0])], context=context)		
        if len(attachment_id)<1: return False

        attach = attachment_obj.browse(cr,uid,attachment_id[0],context=context)
        fname = attach.store_fname
        display_name = attach.name		
        if not fname : return False		
        fname = self.get_full_path(cr, uid, fname)
        wb = load_workbook(filename=fname)	
        ws = wb.get_sheet_by_name(wb.get_sheet_names()[0])
        highest_row = ws.get_highest_row()
        highest_col = ws.get_highest_column()
        title_name = ws.cell(row = 0,column = 0).value	
        title_quantity = ws.cell(row = 0,column = 1).value		

        if highest_col < 2 or title_name != "name" or title_quantity != "quantity":	            		
            raise osv.except_osv(u'Excel错误',u'''文件：%s 格式不正确.''' % display_name)	

        # 读取 Excel 文件内的行到列表中， 格式必须为 2列(name,quantity)			
        row_start = 1
        lines = []	
        product_obj = self.pool.get('product.product')		
        while row_start < highest_row :
            name = ws.cell(row=row_start,column=0).value
            name = name.strip()			
            qty_tmp = ws.cell(row=row_start,column=1)			
            quantity = qty_tmp.get_original_value() or 1
			
            product_ids = product_obj.search(cr, uid, [('name_template','=',name)], context=context)  
            if not product_ids : raise osv.except_osv(u'产品名错误',u'''没有产品： %s 。''' % name)		  # 判断产品名称 是否存在于 ERP中	
            lines.append((product_ids[0],quantity))	
			
            row_start += 1			# 必须加此列，否则死循环

        # 设定 stock_move 的基本数据			
        picking_obj = self.pool.get('stock.picking').browse(cr,uid,ids[0],context=context)
        picking_type = picking_obj.picking_type_id		
        vals = {	
            'product_id': 0,  
            'product_uom_qty':1, 
            'location_dest_id':picking_type.default_location_dest_id.id, 
            'location_id': picking_type.default_location_src_id.id, 
            'company_id': picking_obj.company_id.id, 
            'date':datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            'date_expected':(datetime.datetime.now() + datetime.timedelta(3)).strftime("%Y-%m-%d %H:%M:%S"), 
            'invoice_state':'none', 
            'name':'-', 
            'procure_method':'make_to_stock', 
            'state':'draft',			
            'product_uom':1, 
            'weight_uom_id':1,
            'picking_id': ids[0],			
        }	
		
        # 按照 保存的行列表内容 逐行 创建 Stock Move.		
        move_obj = self.pool.get('stock.move')		
        for line in lines :
            vals_move = vals.copy()
            vals_move.update({'product_id':line[0], 'product_uom_qty':line[1]})			
            move_obj.create(cr, uid, vals_move, context=context)                    		
					
    def split_picking_by_shuadan(self, cr, uid, ids, context=None):
	
        picking_obj = self.pool.get('stock.picking').browse(cr,uid,ids[0],context=context)
        if picking_obj.state in ['done', 'cancel', 'draft'] : return True
		
        statement = "select count(id) from stock_move where picking_id=%d and product_id in (20808,20950)"	% ids[0]
        cr.execute(statement)
        res = 0		
        for i in cr.fetchall():
            res = i[0]		
        if not res : 
            raise osv.except_osv(u'错误',u'''没有 刷单空包产品，不进行拆分！''')
            return True			
		
        new_obj = picking_obj.copy()
        statement = "delete from stock_move where picking_id=%d; update stock_move set picking_id=%d where picking_id=%d and product_id in (20808,20950); update stock_picking set state='assigned',invoice_state='2binvoiced' where id=%d" % (  new_obj.id, new_obj.id, ids[0], new_obj.id )
        cr.execute(statement)				
	
    def import_orders_from_note(self, cr, uid, ids, context=None):	
        picking_obj = self.pool.get('stock.picking').browse(cr,uid,ids[0],context=context)	
        customer_id = picking_obj.partner_id.id	
        carrier_obj = self.pool.get('loewie.carrier')			
        coe_obj = self.pool.get('sale.coe')		
        lines = [ o.split(u'，') for o in picking_obj.note.split(chr(10))]	

        express_ids = []		
        pay_way = {'we_pay':u'包邮' , 'customer_pay': u'到付', 'cash_pay': u'现付'}		
        for line in lines:
            if len(line)<4: 
                lines.remove(line)		
                continue					
				
            vals = {		
                'picking_id': ids[0],
                'customer': customer_id,				
                'name':line[0].strip(),
                'mobile':line[1].strip(), 					
                'address':line[2].strip(),			
                'zip':line[3].strip(),
                'price':6,				
            }
			
            express_id = coe_obj.search(cr, uid, [('name','=',vals["name"]),('mobile','=',vals['mobile']),('picking_id','=',ids[0])], context=context)	
            express_id = express_id and express_id[0] or 0			
            if not express_id :	
								
                way = len(line)>5 and line[5] or ''	
                way = way.strip()				
                way_erp = ''			
                for key in pay_way.keys():
                    if pay_way[key] == way : way_erp = key
                if way_erp in pay_way.keys(): vals.update({'pay_way':way_erp})					
			
                expresser = line[4] or ''
                carrier_id = carrier_obj.search(cr,uid,[('name','=',expresser.strip())],context=context)
                carrier_id = carrier_id and carrier_id[0] or 0			
                if carrier_id : vals.update({'expresser':carrier_id})			
                express_id = coe_obj.create(cr, uid, vals, context=context)
            express_ids.append(express_id)            				
			
        if len(express_ids) != 1:
            return True

        for move in picking_obj.move_lines:
            move.express_id2 = express_ids[0]		
			
        return True
				
    def view_express_data(self, cr, uid, ids, context=None):	
        stock_picking_obj = self.pool.get('stock.picking').browse(cr,uid,ids[0],context=context)	

        mod_obj = self.pool.get('ir.model.data')
        act_obj = self.pool.get('ir.actions.act_window')
        result = mod_obj.get_object_reference(cr, uid, 'loewieec_sync_sz', 'action_loewieec_salecoe')
        id = result and result[1] or False
        result = act_obj.read(cr, uid, [id], context=context)[0]		
        partner_id = stock_picking_obj.partner_id.id
		
        if stock_picking_obj.sale_id and not stock_picking_obj.sale_id.is_set_seq:
            express_ids = []
            shuada_id = self.pool.get('product.product').search(cr,uid,[('name_template','=',u'刷单空包')],context=context)			
            if not shuada_id : return False			
            else: shuada_id = shuada_id[0]			
            for sale_line in stock_picking_obj.sale_id.order_line:
                if not sale_line.express_id or sale_line.express_id in express_ids : continue			
                if sale_line.product_id.id == shuada_id :  			
                    sale_line.express_id.is_shuadan = True
                express_ids.append(sale_line.express_id)					
            		
		
        if not stock_picking_obj.sale_id : 
            domain = [('picking_id','=',ids[0]),('express_id2','!=',False)]		# ,('express_id2','!=',False)
            express_ids = self.pool.get('stock.move').read_group(cr, uid, domain, ['express_id2'], ['express_id2'], context=context)
            express_ids = map(lambda x: (x['express_id2'] and x['express_id2'][0]), express_ids)			
            result['domain'] = [('id','in', express_ids)]
            result['res_id'] = express_ids
            result['context'] = {'default_customer':partner_id,'default_picking_id':ids[0]}				
            return result
			
        order_id = 	stock_picking_obj.sale_id.id 		
        #sale_order_line_ids = self.pool.get('sale.order.line').search(cr,uid,[('order_id','=',order_id)],context=context)	
        #if len(sale_order_line_ids)< 1: return False		
        #eids = self.pool.get('sale.order.line').read(cr,uid,sale_order_line_ids,['express_id'],context=context)
        #express_ids = [ eid['express_id'] and eid['express_id'][0] for eid in eids ]			
        #if len(express_ids) < 1: 			
        #    raise osv.except_osv(u'stock.picking 错误',u'''没有快递信息''')		
        #    return False	
		
        express_ids = []
        for move in stock_picking_obj.move_lines:
            express = move.procurement_id.sale_line_id.express_id		
            tmp_id = express and express.id
            if tmp_id not in express_ids : express_ids.append(tmp_id)			
	
        sale_coe_obj = self.pool.get('sale.coe')
        for express_obj in sale_coe_obj.browse(cr,uid,express_ids,context=context):
            if not express_obj.picking_id: 
                express_obj.picking_id = ids[0]				

        result['domain'] = [('id','in', express_ids)]
        result['res_id'] = express_ids
        result['context'] = {'default_sale_id':order_id,'default_customer':partner_id,'default_picking_id':ids[0]}	
		
        return result		
	
class sale_order(osv.osv):
    _name = "sale.order"
    _inherit = "sale.order"
        
    _columns = {
        'express_ids': fields.related('order_line', 'express_id', type='many2one', relation='sale.coe', string=u'TMI_JDI收货人'),	
        'tmi_jdi_nos': fields.related('order_line', 'tmi_jdi_no', type='char', string='TMI_JDI_NO'),		
        'is_set_seq': fields.boolean('Set Sequence',default=False, readonly=True),		
        'selected': fields.boolean('Selected'),	
        'shop_id': fields.many2one('loewieec.shop', string=u"EC店铺名", readonly=True),
        'sale_code': fields.char(u'EC单号', readonly=True),				
        'order_state': fields.selection([
            ('WAIT_SELLER_SEND_GOODS', u'等待卖家发货'),
            ('WAIT_BUYER_CONFIRM_GOODS', u'等待买家确认收货'),
            ('TRADE_FINISHED', u'交易成功'),
            ('TRADE_CLOSED', u'交易关闭'),
            ], u'订单状态'),
    }

    def update_orders_seller_memo(self, cr, uid, ids, context=None):
        sale_order_obj = self.pool.get('sale.order').browse(cr,uid,ids[0],context=context)	
        shop = sale_order_obj.shop_id		
        if not shop : return False 
        if shop.code == 'JDI' :
            raise osv.except_osv(u'错误',u'''JDI京东国际订单无需更新备注''') 
            return False			
			
        statement = "select tmi_jdi_no from sale_order_line where order_id=%d group by tmi_jdi_no" % ids[0] 
        cr.execute(statement)
        tids = [item[0] for item in cr.fetchall()]
        if not tids : return False
		
        return shop.update_orders_seller_memo(context=context, tids=tids)	
	
    def delete_lines_of_tmijdi_no(self, cr, uid, ids, context=None):   # 完整删除 天猫京东 订单的 行
        sale_order_obj = self.pool.get('sale.order').browse(cr,uid,ids[0],context=context)

        note = 	sale_order_obj.note or ''
        tmijdi_nos = note.strip().split(',') 		
        tmijdi_no_list = []
	
        for tmijdi_no in tmijdi_nos:
            if tmijdi_no.strip() != '': tmijdi_no_list.append( tmijdi_no.strip() )
			
        statement = "delete from sale_order_line where order_id=%d and tmi_jdi_no in (%s)" %  ( ids[0], ("'" + """','""".join(tmijdi_no_list)	+ "'") )		
        cr.execute(statement)				

        val = val1 = 0.0
        cur = sale_order_obj.pricelist_id.currency_id
        for line in sale_order_obj.order_line:
            val1 += line.price_subtotal
            val += self._amount_line_tax(cr, uid, line, context=context)
			
        cur_obj = self.pool.get('res.currency')			
        amount_tax = cur_obj.round(cr, uid, cur, val)		
        amount_untaxed = cur_obj.round(cr, uid, cur, val1)		
        amount_total = amount_untaxed + amount_tax		
        sale_order_obj.write({'amount_tax':amount_tax, 'amount_untaxed': amount_untaxed,'amount_total':amount_total})	
	
    def update_waybill_no(self, cr, uid, ids, context=None):
        sale_order_obj = self.pool.get('sale.order').browse(cr,uid,ids[0],context=context)	
        shop = sale_order_obj.shop_id		
        if not shop : return False 
        if shop.code == 'JDI' :
            return shop.jdi_order_delivery(salesorder=sale_order_obj, context=context)  		
        
        return shop.update_tmall_waybill(context=context, salesorder=sale_order_obj)			
		
    def string_refactor(self, name):   
        if not name: return False		
        name = name.replace("/","%")
        name = name.replace("|","%")			
        ll = name.split("-")
        ll = [l.strip() for l in ll]
        ll = "%".join(ll)				
        return ll.replace(" ","%")
		
    def get_express_data(self, cr, uid, ids, context=None):			
		
        sale_coe_obj = self.pool.get('sale.coe')
        express_ids = sale_coe_obj.search(cr,uid,[('sale_id','=',ids[0])],context=context)
        if len(express_ids)<1: return False
		
        note = ""		
        for express in sale_coe_obj.browse(cr,uid,express_ids,context=context):
            if not express.name : continue		
            name = express.name
            express_no = express.express_no or 'none'
            expresser = express.expresser and express.expresser.name or 'none'			
            tmp_str =  name + "," + expresser + "," + express_no + chr(10)		
            note += tmp_str
			
        sale_order_obj = self.pool.get('sale.order').browse(cr,uid,ids[0],context=context)
        note_tmp = sale_order_obj.note	or  '.'	
        sale_order_obj.note = note + note_tmp 
        return True		
		
    def view_express_data(self, cr, uid, ids, context=None):	
        sale_order_obj = self.pool.get('sale.order').browse(cr,uid,ids[0],context=context)		
        if not sale_order_obj : 
            raise osv.except_osv(u'Sale order 错误',u'''请先保存销售单草稿''')	
            return False

        #express_ids = sale_coe_obj.search(cr,uid,[('sale_id','=',ids[0])],context=context)
        sale_order_line_ids = self.pool.get('sale.order.line').search(cr,uid,[('order_id','=',ids[0])],context=context)	
        if len(sale_order_line_ids)< 1: return False		
        eids = self.pool.get('sale.order.line').read(cr,uid,sale_order_line_ids,['express_id'],context=context)
        #if len(eids) < 1: return False		
        express_ids = [ eid['express_id'] and eid['express_id'][0] for eid in eids ]			
		
        customer_id = sale_order_obj.partner_id.id
        sale_coe_obj = self.pool.get('sale.coe')
        if len(express_ids)>0:		
            for express_obj in sale_coe_obj.browse(cr,uid,express_ids,context=context):
                express_obj.sale_id = ids[0]		
                express_obj.customer = customer_id	                		
		
        mod_obj = self.pool.get('ir.model.data')
        act_obj = self.pool.get('ir.actions.act_window')

        result = mod_obj.get_object_reference(cr, uid, 'loewieec_sync_sz', 'action_loewieec_salecoe')
        id = result and result[1] or False
        result = act_obj.read(cr, uid, [id], context=context)[0]
        result['domain'] = [('id','in',express_ids)]
        result['res_id'] = express_ids
        result['context'] = {'default_sale_id':ids[0],'default_customer':customer_id}		
		
        return result		
		
    def import_orders_from_note(self, cr, uid, ids, context=None):	
        sale_order_obj = self.pool.get('sale.order').browse(cr,uid,ids[0],context=context)	
        if sale_order_obj.shop_id or sale_order_obj.partner_id.name in [u'TMI天猫国际', u'天猫乐易成人用品专营店'] :
            return self.import_tmall_csv_file(cr, uid, ids, context=context)	# 2边ERP统一收货人信息 的 导入方式 为 从附件导入	
            #return self.import_tmall_consignee_from_note(cr, uid, ids, context=context, order=sale_order_obj)   				
			
        if not sale_order_obj : 
            raise osv.except_osv(u'Sale order 错误',u'''请先保存销售单草稿''')	
            return False
        customer_id = sale_order_obj.partner_id.id	
        carrier_obj = self.pool.get('loewie.carrier')			
        coe_obj = self.pool.get('sale.coe')	
        note = sale_order_obj.note.strip()		
        lines = [ o.split(u'，') for o in note.split(chr(10))]	

        express_ids = []		
        pay_way = {'we_pay':u'包邮' , 'customer_pay': u'到付', 'cash_pay': u'现付'}		
        for line in lines:
            if len(line) < 4 : continue		
            if len(line) != 6 or not line[3].strip().isdigit() or line[2].strip().isdigit():
                raise osv.except_osv(u'格式错误',','.join(line) + chr(10) + chr(10) + u"正确格式例子：姓名，手机号码，地址，邮编，韵达，现付" )							
				
            vals = {		
                'sale_id': ids[0],
                'customer': customer_id,				
                'name':line[0].strip(),
                'mobile':line[1].strip(), 					
                'address':line[2].strip(),			
                'zip':line[3].strip(),
                'price':6,				
            }		
            express_id = coe_obj.search(cr, uid, [('name','=',vals["name"]),('mobile','=',vals['mobile']),('sale_id','=',ids[0])], context=context)	
            express_id = express_id and express_id[0] or 0			
            if not express_id :	
								
                way = len(line)>5 and line[5] or ''	
                way = way.strip()				
                way_erp = ''			
                for key in pay_way.keys():
                    if pay_way[key] == way : way_erp = key
                if way_erp in pay_way.keys(): vals.update({'pay_way':way_erp})					
			
                expresser = line[4] or ''
                carrier_id = carrier_obj.search(cr,uid,[('name','=',expresser.strip())],context=context)
                carrier_id = carrier_id and carrier_id[0] or 0			
                if carrier_id : vals.update({'expresser':carrier_id})			
                express_id = coe_obj.create(cr, uid, vals, context=context)
            express_ids.append(express_id)            				
			
        if len(express_ids) != 1:
            return True

        for product in sale_order_obj.order_line:
            product.express_id = express_ids[0]		
			
        return True
		
    def import_tmall_consignee_from_note(self, cr, uid, ids, context=None,order=None):			
        sale_order_obj = order		
        if not sale_order_obj : 
            raise osv.except_osv(u'Sale order 错误',u'''请先保存销售单草稿''')	

        customer_id = sale_order_obj.partner_id.id	
        carrier_obj = self.pool.get('loewie.carrier')			
        coe_obj = self.pool.get('sale.coe')		
        lines = [ o.split(chr(9)) for o in sale_order_obj.note.split(chr(10))]	
		
        statement = "select tmi_jdi_no from sale_order_line where order_id=%d group by tmi_jdi_no" % ids[0]		
        cr.execute(statement)
        tids = [item[0] for item in cr.fetchall()]
        lost_tids = []		
		
        express_ids = {}		
        pay_way = {'we_pay':u'包邮' , 'customer_pay': u'到付', 'cash_pay': u'现付'}		
        for line in lines:
            if len(line) < 4 : continue	
            tmino = line[0].strip()			
            if len(line) != 4 or not tmino.isdigit() or line[1].strip().isdigit() or line[2].strip().isdigit():
                raise osv.except_osv(u'格式错误',','.join(line) + chr(10) + chr(10) + u'正确格式：订单编号，收件人姓名，收件地址，收件人手机号码')	
				
            if tmino not in tids :  # 如果销售订单行中 天猫单号不在
                lost_tids.append(tmino)
                continue
				
            vals = {		
                'sale_id': ids[0],
                'customer': customer_id,
                'tmi_jdi_no': tmino,			
                'name':line[1].strip(),	
                'mobile':line[3].strip(),				
                'address':line[2].strip(), 	
                #'zip':line[4].strip(),
                'expresser':3,
                'pay_way':'cash_pay',				
                'price':6,				
            }		
            express_id = coe_obj.search(cr, uid, [('name','=',vals["name"]),('mobile','=',vals['mobile']),('sale_id','=',ids[0])], context=context)	
            express_id = express_id and express_id[0] or 0			
            if not express_id :			
                express_id = coe_obj.create(cr, uid, vals, context=context)
            express_ids[vals['tmi_jdi_no']] = express_id            			

        for product in sale_order_obj.order_line:
            tmi_no = product.tmi_jdi_no.strip()		
            if tmi_no in express_ids.keys():		
                product.express_id = express_ids[tmi_no]	

        if lost_tids :
            note = sale_order_obj.note	or  '-'	
            note_tmp = u"销售单中不存在以下电商单号:" + ",".join(lost_tids) + chr(10)			
            sale_order_obj.note = note_tmp + note 		
			
        return True
				
    def set_line_express_id(self, cr, uid, ids, context=None):		
        sale_order_obj = self.pool.get('sale.order').browse(cr,uid,ids[0],context=context)	
        express_ids = self.pool.get('sale.coe').search(cr,uid,[('sale_id','=',ids[0])],context=context)
        express_id = len(express_ids)==1 and express_ids[0]	or 0
        if not express_id : return False		
		
        for product in sale_order_obj.order_line:
            product.express_id = express_id		
			
        return True

    def split_address(self, address):
        assert address != None
        addr_list = address.split(" ",2)		
        if len(addr_list) < 3: return None	
		
        length = len( addr_list[2] )
        for i in range( length-1, 0, -1 ):
            if addr_list[2][i] == '(': break

        addr = addr_list[2][0:i]
        zip = addr_list[2][i+1: length-1]		
		
        return {'province':addr_list[0],'city':addr_list[1],'address':addr,'zip':zip}				
		
    def get_full_path(self, cr, uid, path):
        # sanitize ath
        path = re.sub('[.]', '', path)
        path = path.strip('/\\')
        return os.path.join(tools.config.filestore(cr.dbname), path)	
		
    def import_tmall_csv_file(self, cr, uid, ids, context=None):

        attachment_obj = self.pool.get('ir.attachment')
        shop_attach_id = context.get('res_id')	
        shop_attach_id = shop_attach_id or ids[0]		
        attachment_id = attachment_obj.search(cr,uid,[('res_id', '=', shop_attach_id)], context=context)		
        if len(attachment_id)<1: return False
        attach = attachment_obj.browse(cr,uid,attachment_id[0],context=context)
        fname = attach.store_fname
        display_name = attach.name		
        if not fname : return False		
        fname = self.get_full_path(cr, uid, fname)	
		
        csvfile = file(fname, 'rb')		
		
        coe_obj = self.pool.get('sale.coe')    		
        tmi_no = {}
        no_coe_list = []		
        		
        statement = 'select tmi_jdi_no from sale_order_line where order_id=%d group by tmi_jdi_no' % ids[0]
        cr.execute(statement)
        tmi_jdi_no_list = [ str(item[0]).strip() for item in cr.fetchall() ]		
				
        sale_id = ids[0]
        #for line in reader:
        for cols in csvfile.readlines():		
            cols = cols.decode('gbk')
            cols = cols.split(",")	
            line = [ col.replace('"','').strip() for col in cols ]
            line[0] = line[0].replace('=','').strip()			
            if line[0] == u'订单编号' and line[1] == u'买家会员名' or line[0] not in tmi_jdi_no_list :	 continue

            #coeids = coe_obj.search(cr,uid,[('name','=',line[12]),('address','=',addr and addr2 or line[13]),('tel','=',phone)],context=context)
            coeids = coe_obj.search(cr,uid,[('tmi_jdi_no','=',line[0])],context=context)			
            coeid = coeids and coeids[0] or 0				
            if not coeid : 	
			
                addr = self.split_address(line[13])		
                addr2 = addr['province'] + ',' + addr['city'] + ',' + addr['address']
                phone = line[16].replace("'","")
                phone = phone.strip()
			
                if addr : coe_info = {'tmi_jdi_no':line[0], 'name':line[12], 'receive_name':line[12], 'mobile':phone, 'tel':phone, 'address':addr2, 'province':addr['province'], 'city':addr['city'], 'zip':addr['zip']}
                else: coe_info = {'tmi_jdi_no':line[0], 'name':line[12], 'receive_name':line[12], 'mobile':phone, 'tel':phone, 'address':line[13]}				
                coeid = coe_obj.create( cr, uid, coe_info, context=context )
 
            tmi_no[line[0]] = coeid
			
        sale_order = self.pool.get('sale.order').browse(cr,uid,ids[0],context=context)
        coe_in_sale = []		
        for line in sale_order.order_line:
            tmi_jdi_no = line.tmi_jdi_no and line.tmi_jdi_no.strip() or ''
            if not tmi_jdi_no : continue
			
            if tmi_jdi_no in tmi_no.keys():
                line.write( {'express_id': tmi_no[tmi_jdi_no]} )  # 深圳ERP中为 line.express_id 而香港ERP为 line.coe_no
                coe_in_sale.append(tmi_jdi_no)	
                #_logger.info( "Jimmy: have such a order:%s, and coeid:%d " % ( tmi_jdi_no, tmi_no[tmi_jdi_no]) )				
                #tmi_no[tmi_jdi_no] = 0	 # 这里有问题， 如果一个 电商订单有几个产品，则只有第一个产品行 会被添加 coe 条目id			
            else:			
                no_coe_list.append(tmi_jdi_no) 

        if no_coe_list : 
            log = sale_order.note or ''		
            sale_order.note = u'以下销售订单行的TMI_NO在CSV文件中不存在:' + chr(10) + ",".join(no_coe_list) + chr(10) + log
			
        not_in_tmi_no = []
        for key in tmi_no.keys():
            if key not in  coe_in_sale: not_in_tmi_no.append(key)
			
        if not_in_tmi_no : 
            log = sale_order.note or ''		
            sale_order.note = u"以下CSV内的 '订单编号' 无法匹配到 销售订单行:" + chr(10) + ",".join( not_in_tmi_no ) # + chr(10) + log	  #
			
        #_logger.info( "Jimmy: OK....")			
        return True			