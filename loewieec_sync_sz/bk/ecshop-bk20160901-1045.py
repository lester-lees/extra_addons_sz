# -*- encoding: utf-8 -*-
import time
import logging
from openerp import tools
from dateutil.relativedelta import relativedelta
from datetime import datetime, timedelta
from openerp.tools.translate import _
from openerp.osv import fields,osv
import json
import hashlib
import chardet
import os
import re

from top import setDefaultAppInfo
from top.api.rest import ItemsOnsaleGetRequest
from top.api.rest import TradesSoldIncrementGetRequest
from top.api.rest import TimeGetRequest
from top.api.rest import TradesSoldGetRequest
from top.api.rest import TradeFullinfoGetRequest
from top.api.rest import ItemSkusGetRequest
from top.api.rest import ItemQuantityUpdateRequest
from top.api.rest import SkusQuantityUpdateRequest
from top.api.rest import ItemsInventoryGetRequest
from top.api.rest import ProductsSearchRequest
from top.api.rest import WlbInventorySyncRequest

from openpyxl.reader.excel import load_workbook 
from openerp import tools

_logger = logging.getLogger(__name__)

class loewieec_shop(osv.osv):
    _name = 'loewieec.shop'
    _description = u"电商店铺"
   
    _columns = {
        'tmall_time': fields.datetime(string='TMall Time'),	
        'name': fields.char(u'店铺名称', size=16, required=True),
        'code': fields.char(u'店铺前缀', size=8, required=True, help=u"系统会自动给该店铺的订单编号.客户昵称加上此前缀.通常同一个平台的店铺前缀设置成一样"),
        'platform': fields.selection([('tb', u'淘宝天猫'), ('sb', u'淘宝沙箱'),], string=u'电商平台', required=True, help = u"淘宝、京东等电商平台" ),
        #'categ_id': fields.many2one('product.category', string=u"商品默认分类", required=True),
        'location_id': fields.many2one('stock.location', string=u"店铺库位", required=True),
        'journal_id': fields.many2one('account.journal', string=u"默认销售账簿", required=True),
        'post_product_id': fields.many2one('product.product', string=u"邮费"),
        #'coupon_product_id': fields.many2one('product.product', string=u"优惠减款", required=True),
        'gift_product_id': fields.many2one('product.product', string=u"赠品", ),
        'partner_id': fields.many2one('res.partner',string='Partner', required=True),
        'pricelist_id':fields.many2one('product.pricelist',string='Price List', required=True),		
        'warehouse_id':fields.many2one('stock.warehouse',string='Warehouse', required=True),		
        'appkey': fields.char(u'App Key', ),
        'appsecret': fields.char(u'App Secret', ),
        'sessionkey': fields.char(u'Session Key', ),
        'apiurl': fields.char(u'API URL', ),
        'authurl': fields.char(u'Auth URL', ),
        'tokenurl': fields.char(u'Token URL', ),
      	'products': fields.one2many('product.tmalljd','ec_shop_id',string="Products"),  
        'orders': fields.one2many('sale.order','shop_id',string="Orders"),  		
        'start_modified': fields.datetime(string='Start Modified'),	
        'end_modified': fields.datetime(string='End Modified'),	
        'sync_interval': fields.integer(u'同步最近多少小时的订单',default=24),
        'import_salesorder': fields.many2one('sale.order',u'天猫COE信息的目标订单'),	
        'last_log':fields.text(u'同步信息'),	
        'tmi_state': fields.selection([
            ('WAIT_BUYER_PAY', u'等待买家付款'),
            ('WAIT_SELLER_SEND_GOODS', u'等待卖家发货'),
            ('WAIT_BUYER_CONFIRM_GOODS', u'等待买家确认收货'),
            ('TRADE_FINISHED', u'交易成功'),
            ('TRADE_CLOSED', u'交易关闭'),
            ('TRADE_CLOSED_BY_TAOBAO', '交易被淘宝关闭'),
            ('ALL_WAIT_PAY', u'所有买家未付款的交易'),
            ('ALL_CLOSED', u'所有关闭的交易'),
            ], '订单的天猫订单状态', default='WAIT_SELLER_SEND_GOODS' ),		
    }

    def get_full_path(self, cr, uid, path):
        # sanitize ath
        path = re.sub('[.]', '', path)
        path = path.strip('/\\')
        return os.path.join(tools.config.filestore(cr.dbname), path)

    def create_order_from_excel(self, cr, uid, ids, coe_lines, product_lines, tmijdi_fname, context=None):	
        order_obj = self.pool.get('sale.order')
        shop = self.browse(cr, uid, ids[0], context = context)
        product_gift_id = shop.gift_product_id.id		
        partner_id = shop.partner_id.id		
		
        # 检查所有产品名是否 在ERP中有匹配的项		
        product_obj = self.pool.get('product.product')
        for product in product_lines:			
            product_name = 	product.get('product',"").strip()		
            product_ids = product_obj.search(cr, uid, [('name_template','=',product_name)], context = context )
            if not product_ids:
                name = self.string_refactor(product_name)			
                product_ids = product_obj.search(cr, uid, [('name_template','like',name)], context = context )
                if product_ids: product['desc'] = u"产品名未能精确匹配" 
				
            if not product_ids:
                raise osv.except_osv(u'Excel产品名错误',u'''无法再ERP中找到 - %s. \r\n请检查名字 并纠正如下情况：\r\n 1. 125 ml -> 125ml. \r\n 2. Eros - Eros Warming - 150ml  -> Eros - Warming - 150ml''' % product_name)			
                return False
            product['product_id'] = product_ids[0]	
			
        # 销售订单 值列表	
        order_name = tmijdi_fname.strip().split(".")[0]		
        order_val = {
            'name': tmijdi_fname.strip().split(".")[0],
            'shop_id': shop.id,
            'date_order':  datetime.now(),      #订单支付时间
            'partner_id': partner_id,
            'partner_invoice_id': partner_id, 			
            'partner_shipping_id': partner_id,
            'warehouse_id': shop.warehouse_id.id,
            'pricelist_id': shop.pricelist_id.id,
            'company_id': 1,			
            'all_discounts': 0,
            'picking_policy': 'one',
            'state':'draft',		
            'user_id': uid, 			
            'order_policy': 'picking',
            'client_order_ref': order_name,			
            'order_line': [],
            'note': '',			
        }
		
        note = ""			
        coe_obj = self.pool.get('sale.coe')			
        for key in coe_lines:
            coe_line = coe_lines[key]		
            coe_vals = {	# 为每个收件人 创建COE 条目 		
                'tmi_jdi_no': coe_line["tmijdi_no"],
                'name': coe_line["coe_no"],
                'receive_name': coe_line["receive_name"],
                'tel': coe_line["tel"],
                'state': coe_line["state"],
                'city': coe_line["city"],
                'address': coe_line["address"],
                'zip': coe_line["zip"],
            }
            coe_id = coe_obj.search(cr, uid, [('name','=',coe_line["coe_no"])], context=context)
            if coe_id : coe_id = coe_id[0]			
            if not coe_id :			
                coe_id = coe_obj.create(cr, uid, coe_vals, context=context)	
            coe_line["id"] = coe_id	
           
            gift_product_vals = {   # 为每个 TMI/JDI单 加入一个赠品			
                'product_uos_qty': 1,
                'product_id': product_gift_id,
                'product_uom': 1,
                'coe_no': coe_id,				
                'price_unit': 0,
                'product_uom_qty': 1,
                'name': '.',
                'delay': 7,
                'discount': 0,
            }						
            order_val['order_line'].append( (0, 0, gift_product_vals)  )	
		
            note += u"TMJD NO:%s, COE:%s, 姓名:%s, 电话:%s, 省份:%s, 城市%s, 收货地址:%s, 邮编:%s;"	% (coe_line["tmijdi_no"], coe_line["coe_no"], coe_line["receive_name"],coe_line["tel"],coe_line["state"],coe_line["city"],coe_line["address"],coe_line["zip"] )
            note += chr(10)	

        order_val.update({"note":note})    		
		
        for o in product_lines:
            desc = o.get('desc','.').strip()
            if desc == "" : desc = '-'			
            line_vals = {			
                'product_uos_qty': o.get('qty',0),
                'product_id': o['product_id'],
                'product_uom': 1,
                'coe_no': coe_lines[o.get('tmijdi_no')]["id"],				
                'price_unit': o.get('price',0),
                'product_uom_qty': o.get('qty',0),
                'name': desc,
                'delay': 7,
                'discount': 0,
            }				
            order_val['order_line'].append(  (0, 0, line_vals)  ) 
	
        order_id = order_obj.create(cr, uid, order_val, context = context)

        return order_id
		
    def string_refactor(self, name):

        if not name: return False
		
        name = name.replace("/","%")
        name = name.replace("|","%")			
        ll = name.split("-")
        ll = [l.strip() for l in ll]
        ll = "%".join(ll)		
		
        return ll.replace(" ","%")
		
    def import_orders_from_excel(self, cr, uid, ids, context=None):
        attachment_obj = self.pool.get('ir.attachment')
        attachment_id = attachment_obj.search(cr,uid,[('res_id', '=', ids[0])], context=context)		
        if len(attachment_id)<1: return False

        attach = attachment_obj.browse(cr,uid,attachment_id[0],context=context)
        fname = attach.store_fname
        display_name = attach.name		
        if not fname : return False		
        fname = self.get_full_path(cr, uid, fname)
        wb = load_workbook(filename=fname)	
        #ws = wb.get_sheet_by_name("Sheet1")
        ws = wb.get_sheet_by_name(wb.get_sheet_names()[0])
        highest_row = ws.get_highest_row()
        highest_col = ws.get_highest_column()
        title_order = ws.cell(row = 1,column = 1).value	
        title_product = ws.cell(row = 1,column = 2).value
        title_qty = ws.cell(row = 1,column = 3).value
        title_coe = ws.cell(row = 1,column = 5).value		

        if highest_col < 12 or title_order != u"订单編號" or title_product != u"产品名称" or title_coe != u"COE" :
            #attach.unlink()	            		
            raise osv.except_osv(u'Excel错误',u'''文件：%s 格式不正确.''' % display_name)		
			
        row_start = 2
        coe_lines = {}	
        last_tmijdi_no = None
        last_coe_no = None		
        while row_start <= highest_row and ws.cell(row=row_start,column=2).value :		
            tmijdi_no = ws.cell(row=row_start,column=1).value
            coe_no = ws.cell(row=row_start,column=5).value	
            name = ws.cell(row=row_start,column=6).value
            tel = ws.cell(row=row_start,column=7).value
            address = ws.cell(row=row_start,column=8).value	
			
            if bool(tmijdi_no) != bool(coe_no) :
                #attach.unlink()	            		
                raise osv.except_osv(u'Excel错误',u'''文件'%s'第：%d行不正确地合并单元格.''' % (display_name, row_start+1) )     
				
            if tmijdi_no and coe_no and name and tel and address: 		
                city = ws.cell(row=row_start,column=9).value			
                state = ws.cell(row=row_start,column=10).value
                zip = str(ws.cell(row=row_start,column=11).value)
                last_tmijdi_no = tmijdi_no = str(tmijdi_no)				
                coe_lines[last_tmijdi_no] = {"id":1, "tmijdi_no": tmijdi_no, "coe_no":coe_no, "receive_name":name, "tel":tel, "address":address, "city":city, "state":state, "zip":zip}				
				
            row_start += 1

        row_start = 2
        product_lines = []
        last_tmijdi_no = None
        last_coe_no = None		
        while row_start <= highest_row and ws.cell(row=row_start,column=2).value :
            line = {}
            line["product_id"] = 1
            line["desc"] = ""			
            line["product"] = ws.cell(row=row_start,column=2).value
            qty_tmp = ws.cell(row=row_start,column=3)			
            line["qty"] = qty_tmp.get_original_value() or 1
            cell_price = ws.cell(row=row_start,column=4)
            price_tmp = cell_price.get_original_value() or 0
            if line["qty"] > 1 : price_tmp = price_tmp/line["qty"]			
            line["price"] = price_tmp
			
			
            tmijdi_no = ws.cell(row=row_start,column=1).value
            coe_no = ws.cell(row=row_start,column=5).value			
            if tmijdi_no : 				
                last_tmijdi_no = str(tmijdi_no)
                last_coe_no = coe_no			

            line['tmijdi_no'] = last_tmijdi_no
            line['coe_no'] = last_coe_no
			
            product_lines.append(line)					
            row_start += 1				
	
        self.create_order_from_excel(cr, uid, ids, coe_lines, product_lines, display_name, context=context)			
        attach.unlink()		
        return True		
	       
    def search_product_sku(self, cr, uid, ids, num_iids=None, context=None):

        port = 80
        shop = self.browse(cr, uid, ids[0], context = context)
        setDefaultAppInfo(shop.appkey, shop.appsecret)
        req = ItemSkusGetRequest(shop.apiurl,port)
        req.fields="sku_id, num_iid, properties, price, status, memo, properties_name, outer_id,quantity,barcode,change_prop,sku_spec_id,extra_id"
        
        if not num_iids : 
            req.num_iids = shop.tokenurl			
        else: 
            req.num_iids = num_iids
			
        resp= req.getResponse(shop.sessionkey)
        skus = resp.get('item_skus_get_response').get('skus', False).get('sku',False)

        return skus

    def create_product_tmalljd_nosku(self, cr, uid, ids, item, context=None):
        product_tmalljd_objs = self.pool.get('product.tmalljd') 	
        product_vals = {'ec_shop_id':ids[0],'ec_price':float(item.get('price')),'ec_qty':item.get('num'),'ec_outer_code':item.get('outer_id'), 'ec_num_iid':str(item.get('num_iid')),'ec_title':item.get('title') } 	
		
        pids = product_tmalljd_objs.search( cr, uid, [('ec_num_iid', '=', str(item.get('num_iid'))),('ec_shop_id','=',ids[0])], context=context)		
        if len(pids) < 1: 			
                product_tmalljd_objs.create(cr, uid, product_vals)  
        return True				
		
    def search_product(self, cr, uid, ids, context=None ):	
        this = self.browse(cr, uid, ids, context=context)[0]	
        res = self._search_product(cr, uid, ids, product_name = None, start_modified=this.start_modified, end_modified=this.end_modified, context=context)

        if len(res) < 1: return 
        product_tmalljd_objs = self.pool.get('product.tmalljd') 
        num_iids = [str(o['num_iid']) for o in res]
        titles = {}
        for item in res:
            titles[str(item.get('num_iid'))] = item.get('title') 
			
        skus_list = []	
        skus_list2 = []		
        interval = 40
        start = end = 0
        final_end = len(num_iids)
		
        while start < final_end:
            if start + interval > final_end	: end = final_end
            else: end = start + interval	
            sub_list = num_iids[start:end]			
            sub_num_iids = ",".join(sub_list)			
            skus_list +=  self.search_product_sku(cr, uid, ids, num_iids=sub_num_iids, context=context)            			
            start = end	
			
        for sku_item in skus_list:
            str_num_iid = str(sku_item.get('num_iid'))		
            if str_num_iid not in skus_list2:
                skus_list2.append( str_num_iid )		
				
        no_sku_num_iids = []		
        for item in res:
            if str(item["num_iid"]) not in skus_list2: 
                self.create_product_tmalljd_nosku(cr,uid,ids,item,context=context)			
			
        for sku in skus_list:
		
            product_vals = {'ec_shop_id':this.id,'ec_price':float(sku.get('price')),'ec_qty':sku.get('quantity'),'ec_outer_code':sku.get('outer_id'),'ec_sku_id':str(sku.get('sku_id')) } #, 'ec_ean13':sku.get('barcode','')}	
            num_iid = str(sku['num_iid'])			
            product_vals['ec_num_iid'] = num_iid	
            product_vals['ec_title'] = titles[num_iid]			
            pids = product_tmalljd_objs.search( cr, uid, [('ec_sku_id', '=', str(sku['sku_id'])),('ec_shop_id','=',this.id)], context=context)		
            #if len(pids)>0: 
            #    product_tmalljd_objs.write(cr,uid,pids[0],product_vals)	            				
            #else:
            if len(pids) < 1: 			
                product_tmalljd_objs.create(cr, uid, product_vals)	
			
        return True	

    def _search_product(self, cr, uid, ids, product_name = None, start_modified = None, end_modified = None, context=None):
        """
        1) 按商品名称，商品修改时间搜索店铺商品
        2) start_modified、end_modified 都是UTC时间，需要加上8小时传给电商平台
        """
        shop_id = self.browse(cr, uid, ids[0], context= context)
        setDefaultAppInfo(shop_id.appkey, shop_id.appsecret)
        req = ItemsOnsaleGetRequest(shop_id.apiurl, 80)
        req.fields="approve_status,num_iid,title,nick,type,num,list_time,price,modified,delist_time,outer_id"		
        if product_name:
            req.q = product_name
        
        req.page_no = 1
        req.page_size = 100
        total_get = 0
        total_results = 100
        res = []
        while total_get < total_results:
            resp= req.getResponse(shop_id.sessionkey)
            total_results = resp.get('items_onsale_get_response').get('total_results')
		
            if total_results > 0:
                res += resp.get('items_onsale_get_response').get('items').get('item')
            total_get += req.page_size			
            req.page_no = req.page_no + 1
        #
        # 时间需要减去8小时
        for r in res:
            r['modified'] = (datetime.strptime(r['modified'],'%Y-%m-%d %H:%M:%S',) - timedelta(hours=8)).strftime('%Y-%m-%d %H:%M:%S')
        return res

    def create_shipping_address(self, cr, uid, partner_id, vals, context=None):
	
        partner_objs = self.pool.get('res.partner')	
        #state_objs = self.pool.get('res.country.state')
        state_id = 0		
        if 	vals.get('state', False) :	
            state_id = self.pool.get('res.country.state').search(cr,uid,[('name','like',vals.get('state'))])
		
        if not partner_id : return 0		
        partner_vals = {
            'parent_id': partner_id,		
            'name': vals.get('name'),
            'phone': vals.get('phone'),
            'street': vals.get('street'),
            'mobile': vals.get('mobile'),			
            'city': vals.get('city'),	
            'zip': vals.get('zip'),				
            'type': 'delivery',			
            'is_company': False,
            'customer': False,
            'supplier': False,
            }  
			
        if state_id : 
            partner_vals.update({'state':state_id[0]})  
			
        id = partner_objs.create(cr, uid, partner_vals, context = context )			
        return id or 0			

    def remove_duplicate_orders(self, cr, uid, orders, context=None):
        sale_obj = self.pool.get('sale.order')
        submitted_references = [o['sale_code'] for o in orders]
        existing_order_ids = sale_obj.search(cr, uid, [('name', 'in', submitted_references)], context = context)
        existing_orders = sale_obj.read(cr, uid, existing_order_ids, ['name'], context=context)
        existing_references = set([o['name'] for o in existing_orders])
        orders_to_save = [o for o in orders if o['sale_code'] not in existing_references]
        return orders_to_save
        
    def search_orders_by_modified_time(self, cr, uid, ids, status = 'WAIT_SELLER_SEND_GOODS', date_start = None, date_end = None, context=None):
        port = 80
        shop = self.browse(cr, uid, ids[0], context = context)
        partner_id = shop.partner_id.id		
        setDefaultAppInfo(shop.appkey, shop.appsecret)
        req = TradesSoldIncrementGetRequest(shop.apiurl,port)
        req.fields="type,seller_nick,buyer_nick,created,sid,tid,status,buyer_memo,seller_memo,payment,discount_fee,adjust_fee,post_fee,total_fee, pay_time,end_time,modified,received_payment,price,alipay_id,receiver_name,receiver_state,receiver_city,receiver_district,receiver_address, receiver_zip,receiver_mobile,receiver_phone,orders.price,orders.num,orders.iid,orders.num_iid,orders.sku_id,orders.refund_status,orders.status,orders.oid, orders.total_fee,orders.payment,orders.discount_fee,orders.adjust_fee,orders.sku_properties_name,orders.outer_iid,orders.outer_sku_id"
        req.status = status
        req.type = "instant_trade,auto_delivery,guarantee_trade,tmall_i18n"		
		
        if shop.start_modified:
            req.start_modified = shop.start_modified			
        if shop.end_modified:
            req.end_modified = shop.end_modified			
        
        res = []
        req.page_no = 1
        req.page_size = 100

        total_get = 0
        total_results = 100
        while total_get < total_results:
            resp= req.getResponse(shop.sessionkey)
            trades = resp.get('trades_sold_increment_get_response').get('trades', False)
            total_results = resp.get('trades_sold_increment_get_response').get('total_results')
            _logger.info("Jimmy total_results :%d" % total_results)				
            if total_results > 0:
                res += trades.get('trade')
            total_get += req.page_size
            req.page_no = req.page_no + 1
            _logger.info("Jimmy page_size :%d" % req.page_size)
            _logger.info("Jimmy total_get :%d" % total_get)	
        # 时间需要减去8小时
        # 单号加上店铺前缀
        order_ids = []
        for trade in res:
            trade['created'] = (datetime.strptime(trade['created'], '%Y-%m-%d %H:%M:%S',) - timedelta(hours=8)).strftime('%Y-%m-%d %H:%M:%S')
            trade['pay_time'] = (datetime.strptime(trade.get('pay_time'), '%Y-%m-%d %H:%M:%S',) - timedelta(hours=8)).strftime('%Y-%m-%d %H:%M:%S')
            trade['sale_code'] = '%s_%s' % (shop.code, trade.get('tid'))
        
        orders = self.remove_duplicate_orders(cr, uid, res, context=context)
        #orders = res		
        for trade in orders:
            #try:
            #    #_logger.info("Jimmy before create_order")		
                #order_id = self.create_order(cr, uid, ids, trade, context = context )			
            order_id = self.create_order(cr, uid, ids, trade, context = context )
            #    _logger.info("Jimmy after create_order")				
            #    order_ids.append(order_id)
            #except Exception, e:
            #    syncerr = u"店铺[%s]订单[%s]同步错误: %s" % (shop.name, trade['tid'], e)
            #    self.pool.get('loewieec.error').create(cr, uid, {'name':syncerr, 'shop_id': shop.id}, context = context )
            #    continue
        return order_ids

    def get_tmall_time(self, cr, uid, ids, context=None ):	
        shop = self.browse(cr, uid, ids[0], context = context)	
        setDefaultAppInfo(shop.appkey, shop.appsecret)
        req = TimeGetRequest(shop.apiurl,80)	
        resp= req.getResponse(shop.sessionkey)
        ttime = resp.get('time_get_response').get('time',False)
		
        if ttime :
            shop.tmall_time = ttime
        return True			
		
    def create_order(self, cr, uid, ids, trade, context=None):	
        order_obj = self.pool.get('sale.order')
        shop = self.browse(cr, uid, ids[0], context = context)	
        partner_id = shop.partner_id.id		
        gift_product_id = shop.gift_product_id.id		
		
        order_val = {
            'name': "%s_%s" % ( shop.code, self.pool.get('ir.sequence').get(cr, uid, 'sale.order') or '/' ),
            'shop_id': shop.id,
            'date_order':  datetime.now(),      #订单支付时间 trade.get('pay_time') or 
            'create_date': datetime.now(),  #trade.get('created'),       #订单创建时间
            'partner_id': partner_id,
            'partner_invoice_id': partner_id, 			
            'partner_shipping_id': partner_id,
            'warehouse_id': shop.warehouse_id.id,
            'pricelist_id': shop.pricelist_id.id,
            'company_id': 1,			
            'all_discounts': 0,
            'picking_policy': 'one',
            'state':'draft',		
            'user_id': uid, 			
            'order_policy': 'picking',
            'client_order_ref': u'Loewieec_sync Generated',			
            'order_line': [],
        }	
        # store_code = order['orders']		 and str(order["store_code"]).strip() == "loewie"		
        order_lines = []
        tids = []	
        last_log = []		
        for order in trade:            
            lines = order["orders"]["order"]			
            for line in lines:
                #line['tid'] = str( order.get('tid') or order.get('oid') )
                store_code = line.get('store_code') or ''				
                if store_code in ['loewie','']	:	
                    line['description'] = "."				
                    if store_code == '' : line['description'] = u'无法确定是否为保税仓订单'				
                    order_lines.append(line)
                    tids.append( str(order['tid']) )					
                else:
                    last_log.append( str(line.get('tid') or line.get('oid')) )					
				
        for order_line in order_lines:  
            tid = str ( order_line.get('tid')	or order_line.get('oid') )	
            line_vals = {			
                'product_uos_qty': order_line.get('num'),
                'product_id': 1,
                'tmi_jdi_no': tid,			
                'product_uom': 1,
                'price_unit': 1,
                'product_uom_qty': order_line.get('num'),
                'name':order_line.get('description'),
                'delay': 7,
                'discount': 0,
            }			
            product_tmalljd_ids = self.pool.get('product.tmalljd').search(cr, uid, [('ec_sku_id','=',order_line.get('sku_id') or order_line.get('num_iid'))], context = context )
            
            #如果没有匹配到产品，报同步异常  coe_lines
            if not product_tmalljd_ids:
                syncerr = u"订单导入错误: 匹配不到商品。tid=%s, 商品num_iid=%s, outer_sku_id=%s, sku_id=%s " % ( tid, order_line.get('num_iid', ''),  order_line.get('outer_sku_id', ''), order_line.get('sku_id', '') )
                self.pool.get('loewieec.error').create(cr, uid, {'name':syncerr, 'shop_id':shop.id }, context = context )
                return False

            product_tmalljd_obj = self.pool.get('product.tmalljd').browse(cr, uid, product_tmalljd_ids[0], context = context)
            product_id = product_tmalljd_obj.erp_product_id.id
            uom_id = product_tmalljd_obj.erp_product_id.uom_id.id			
            #添加订单明细行
            price_unit = float(order_line.get('payment',0))/order_line.get('num') or float(order_line.get('total_fee',0))/order_line.get('num')			
            line_vals.update({'product_id': product_id , 'price_unit':price_unit, 'product_uom': uom_id } )
            order_val['order_line'].append( (0, 0, line_vals) ) 

        for tid in tids:
            gift_vals = {			
                'product_uos_qty': 1,
                'product_id': gift_product_id,
                'tmi_jdi_no': tid,			
                'product_uom': 1,
                'price_unit': 0,
                'product_uom_qty': 1,
                'name': '.',
                'delay': 7,
                'discount': 0,
            }		
		
            order_val['order_line'].append( (0, 0, gift_vals) )	
			
        order_id = order_obj.create(cr, uid, order_val, context = context)
        history_log = shop.last_log or ""		
        if len(last_log) > 0:
            shop.last_log = history_log + u"以下电商单从保税仓发货,so未导入：" + chr(10) + ",".join(last_log)		

        return order_id

    def search_orders_sent_on_tmall(self, cr, uid, ids, context=None):
        return self.search_orders_by_created_time(cr, uid, ids, context=context, status = ['WAIT_BUYER_CONFIRM_GOODS','WAIT_SELLER_SEND_GOODS'])	
		
    def search_orders_by_created_time(self, cr, uid, ids, context=None, status = []):
        port = 80
        shop = self.browse(cr, uid, ids[0], context = context)
        partner_id = shop.partner_id.id		
        setDefaultAppInfo(shop.appkey, shop.appsecret)
        req = TradesSoldGetRequest(shop.apiurl,port) 
		
        #type,created,payment,,status,total_fee,pay_time,end_time,modified,price,alipay_id,seller_nick,buyer_nick,received_payment, discount_fee,adjust_fee,post_fee,receiver_name,receiver_state,receiver_city,receiver_district,receiver_address, receiver_zip,receiver_mobile,receiver_phone,orders.sku_properties_name,orders.refund_status,buyer_memo,seller_memo,orders.iid,orders.adjust_fee,orders.status,orders.discount_fee,orders.price,orders.outer_iid,  WAIT_SELLER_SEND_GOODS
		
        req.fields="sid,tid,orders.num,orders.store_code,orders.num_iid,orders.sku_id,orders.oid, orders.total_fee,orders.payment,orders.outer_sku_id"
        req.type = "instant_trade,auto_delivery,guarantee_trade,tmall_i18n"		
        hours_interval = shop.sync_interval
        if hours_interval : hours_interval -= 8
        else: hours_interval = 16		
        req.start_created = (datetime.now() - timedelta(hours=hours_interval)).strftime('%Y-%m-%d %H:%M:%S')					
        req.end_created = (datetime.now()+ timedelta(hours=8)).strftime('%Y-%m-%d %H:%M:%S')   
		
        res = []
        if len(status) < 1: status.append( shop.tmi_state or 'WAIT_SELLER_SEND_GOODS')
		
        for order_status in status :
            req.status = order_status		
            req.page_no = 1
            req.page_size = 100
            total_get = 0
            total_results = 100
            while total_get < total_results:
                resp= req.getResponse(shop.sessionkey)
                trades = resp.get('trades_sold_get_response').get('trades', False)
                total_results = resp.get('trades_sold_get_response').get('total_results')
                _logger.info("Jimmy total_results :%d" % total_results)				
                if total_results > 0:
                    res += trades.get('trade')
                total_get += req.page_size
                req.page_no = req.page_no + 1
	
        if total_results < 1:
            shop.last_log = u"在%d小时内没有状态：%s 的订单可下载" % (shop.sync_interval, shop.tmi_state)
            return True			
	
        orders = self.remove_duplicate_tmijdi_no(cr, uid, ids, res, shop, context=context)		
        order_id = self.create_order(cr, uid, ids, orders, context = context )
        shop.import_salesorder = order_id
        return order_id

    def remove_duplicate_tmijdi_no(self, cr, uid, ids, orders, shop, context=None):
        statement = "select tmi_jdi_no from sale_order_line where state not in ('cancel','draft') group by tmi_jdi_no"
        cr.execute(statement)
        exist_tmijdi_no = []
        res = []	
        last_log = []
		
        for item in cr.fetchall():
        	exist_tmijdi_no.append(item[0])
      			
        for order in orders:            
            if str(order["tid"]) not in exist_tmijdi_no:
                res.append(order)
            else:
                last_log.append( str(order["tid"]) )
        if len(last_log) > 0:
            shop.last_log = u"以下电商单号ERP中已存在,所以此次未导入: " + chr(10) + ",".join(last_log) + chr(10)
        return res			
		
    def search_orders_by_tid(self, cr, uid, ids, context=None):
        port = 80
        shop = self.browse(cr, uid, ids[0], context = context)
        partner_id = shop.partner_id.id		
        setDefaultAppInfo(shop.appkey, shop.appsecret)
        req = TradeFullinfoGetRequest(shop.apiurl,port)
        #req.fields="type,seller_nick,buyer_nick,created,sid,tid,status,buyer_memo,seller_memo,payment,discount_fee,adjust_fee,post_fee,total_fee, pay_time,end_time,modified,received_payment,price,alipay_id,receiver_name,receiver_state,receiver_city,receiver_district,receiver_address, receiver_zip,receiver_mobile,receiver_phone,orders.price,orders.num,orders.iid,orders.num_iid,orders.sku_id,orders.refund_status,orders.status,orders.oid, orders.total_fee,orders.payment,orders.discount_fee,orders.adjust_fee,orders.sku_properties_name,orders.outer_iid,orders.outer_sku_id"
        req.fields = "type,receiver_name,receiver_state,receiver_city,receiver_district,receiver_address, receiver_zip,receiver_mobile,receiver_phone"
        req.tid = shop.authurl	

        resp= req.getResponse(shop.sessionkey)
        trade = resp.get('trade_fullinfo_get_response').get('trade', False)	
               			
        return True
				
    def import_coe_no(self, cr, uid, ids, context=None):
        attachment_obj = self.pool.get('ir.attachment')
        attachment_id = attachment_obj.search(cr,uid,[('res_id', '=', ids[0])], context=context)		
        if len(attachment_id)<1: return False
        attach = attachment_obj.browse(cr,uid,attachment_id[0],context=context)
        fname = attach.store_fname
        display_name = attach.name		
        if not fname : return False		
        fname = self.get_full_path(cr, uid, fname)
			
        shop = self.browse(cr, uid, ids[0], context = context)
        salesorder = shop.import_salesorder		
        if not salesorder :
            raise osv.except_osv(u'错误',u'''请选择一个目标订单来导入COE收货人信息!!!''')
			
        wb = load_workbook(filename=fname)        
        ws = wb.get_sheet_by_name(wb.get_sheet_names()[0])
        highest_row = ws.get_highest_row()
        highest_col = ws.get_highest_column()
        title_order = ws.cell(row = 0,column = 0).value	
        title_coe = ws.cell(row = 0,column = 1).value		
        title_name = ws.cell(row = 0,column = 2).value

        if highest_col < 10 or title_order != u"參考編號" or title_coe != u"e特快單號" or title_name != u"收件人姓名" :
            #attach.unlink()	            		
            raise osv.except_osv(u'Excel错误',u'''文件：%s 格式不正确.''' % display_name)		
			
        row_start = 1
        coe_lines = {}
        note = ""		
        while row_start <= highest_row and ws.cell(row=row_start,column=0).value :		
            tmijdi_no = str(ws.cell(row=row_start,column=0).value)
            coe_no = ws.cell(row=row_start,column=1).value	 		
            receive_name = ws.cell(row=row_start,column=2).value
            tel = ws.cell(row=row_start,column=3).value
            address = ws.cell(row=row_start,column=4).value
            city = ws.cell(row=row_start,column=6).value			
            state = ws.cell(row=row_start,column=7).value
            zip = str(ws.cell(row=row_start,column=8).value)				
            coe_lines[tmijdi_no] = {"tmi_jdi_no": tmijdi_no, "name":coe_no, "receive_name":receive_name, "tel":tel, "address":address, "city":city, "state":state, "zip":zip}

            note += u"TMJD NO:%s, COE:%s, 姓名:%s, 电话:%s, 省份:%s, 城市%s, 收货地址:%s, 邮编:%s;"	% (tmijdi_no, coe_no, receive_name,tel,state,city,address,zip )
            note += chr(10)				
			
            row_start += 1
			
        # 从ERP数据库中搜索 是否已存在 COE订单号，如未存在则创建新的COE信息，并将ID赋值到 coe_lines的"id" 			
        coe_obj = self.pool.get("sale.coe")			
        for val in coe_lines.values():
            coe_id = coe_obj.search(cr,uid,[('name','=', val["name"])], context=context)    
            coe_id = coe_id and coe_id[0] or 0
            if not coe_id :
                coe_id = coe_obj.create(cr,uid,val,context=context)
            else:
                coe_obj_id = coe_obj.browse(cr,uid,[coe_id], context=context)			
                tmp_val = coe_obj_id.tmi_jdi_no
                if tmp_val.strip() != val["tmi_jdi_no"].strip() :				
                    coe_obj_id.tmi_jdi_no = tmp_val	+ "," + val["tmi_jdi_no"]			
            val["id"] = coe_id				

        # 按照销售订单行的 TMall订单号，找到并填写对应的COE单号。		
        last_log = []			
        for line in salesorder.order_line:
            tmijdino = line.tmi_jdi_no		
            if tmijdino not in coe_lines.keys(): 
                last_log.append(tmijdino)				
                continue		
            line.coe_no = coe_lines[tmijdino]["id"]
			
        # 通知信息			
        if len(last_log)>0:
            shop.last_log = ",".join(last_log) + u" - COE-e特快excel文件中不存在以上TMI单号."		
			
        salesorder.note = note		
        attach.unlink()		
        return True					
		
		