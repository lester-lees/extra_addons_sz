# -*- encoding: utf-8 -*-
import time
import logging
from openerp import tools
from dateutil.relativedelta import relativedelta
from datetime import datetime, timedelta
from openerp.tools.translate import _
from openerp.osv import fields,osv
import json
import hashlib
import chardet
import os
import re

from top import setDefaultAppInfo
from top.api.rest import ItemsOnsaleGetRequest
from top.api.rest import TradesSoldIncrementGetRequest
from top.api.rest import TimeGetRequest
from top.api.rest import TradesSoldGetRequest
from top.api.rest import TradeFullinfoGetRequest
from top.api.rest import ItemSkusGetRequest
from top.api.rest import ItemQuantityUpdateRequest
from top.api.rest import SkusQuantityUpdateRequest
from top.api.rest import ItemsInventoryGetRequest
from top.api.rest import ProductsSearchRequest
from top.api.rest import WlbInventorySyncRequest

from openpyxl.reader.excel import load_workbook 
from openerp import tools

_logger = logging.getLogger(__name__)

class loewieec_shop(osv.osv):
    _name = 'loewieec.shop'
    _description = u"电商店铺"
   
    _columns = {
        'tmall_time': fields.datetime(string='TMall Time'),	
        'name': fields.char(u'店铺名称', size=16, required=True),
        'code': fields.char(u'店铺前缀', size=8, required=True, help=u"系统会自动给该店铺的订单编号.客户昵称加上此前缀.通常同一个平台的店铺前缀设置成一样"),
        'platform': fields.selection([('tb', u'淘宝天猫'), ('sb', u'淘宝沙箱'),], string=u'电商平台', required=True, help = u"淘宝、京东等电商平台" ),
        #'categ_id': fields.many2one('product.category', string=u"商品默认分类", required=True),
        'location_id': fields.many2one('stock.location', string=u"店铺库位", required=True),
        'journal_id': fields.many2one('account.journal', string=u"默认销售账簿", required=True),
        'post_product_id': fields.many2one('product.product', string=u"邮费"),
        #'coupon_product_id': fields.many2one('product.product', string=u"优惠减款", required=True),
        'gift_product_id': fields.many2one('product.product', string=u"赠品", ),
        'partner_id': fields.many2one('res.partner',string='Partner'),
        'pricelist_id':fields.many2one('product.pricelist',string='Price List'),		
        'warehouse_id':fields.many2one('stock.warehouse',string='Warehouse'),		
        'appkey': fields.char(u'App Key', ),
        'appsecret': fields.char(u'App Secret', ),
        'sessionkey': fields.char(u'Session Key', ),
        'apiurl': fields.char(u'API URL', ),
        'authurl': fields.char(u'Auth URL', ),
        'tokenurl': fields.char(u'Token URL', ),
      	'products': fields.one2many('product.tmalljd','ec_shop_id',string="Products"),  
        'orders': fields.one2many('sale.order','shop_id',string="Orders"),  		
        'start_modified': fields.datetime(string='Start Modified'),	
        'end_modified': fields.datetime(string='End Modified'),		
    }

    def get_full_path(self, cr, uid, path):
        # sanitize ath
        path = re.sub('[.]', '', path)
        path = path.strip('/\\')
        return os.path.join(tools.config.filestore(cr.dbname), path)

    def create_order_from_excel(self, cr, uid, ids, trade, context=None):	
        order_obj = self.pool.get('sale.order')
        shop = self.browse(cr, uid, ids[0], context = context)	
        partner_id = shop.partner_id.id		
		#{"no":order_no, "coe":coe, "name":name, "tel":tel, "address":address, "city":city, "state":state, "zip":zip, "products":[line]}
        shipping_id = self.create_shipping_address(cr, uid, partner_id, {'name':trade["name"], 'phone':trade["tel"], 'street':trade["address"], 'city':trade["city"], 'state':trade["state"], 'zip':trade["zip"]}, context=context)	
        note = u"COE:%s, 姓名:%s, 电话:%s, 省份:%s, 城市%s, 收货地址:%s, 邮编:%s."	% (trade["coe"], trade["name"],trade["tel"],trade["state"],trade["city"],trade["address"],trade["zip"] )	
        order_val = {
            'name': "%s_%s" % (shop.code,  trade.get('no')),
            'shop_id': shop.id,
            'date_order':  datetime.now(),      #订单支付时间
            #'create_date': datetime.now(),       #订单创建时间
            'partner_id': partner_id,
            'partner_invoice_id': partner_id, 			
            'partner_shipping_id': shipping_id,
            'warehouse_id': shop.warehouse_id.id,
            'pricelist_id': shop.pricelist_id.id,
            'company_id': 1,			
            'all_discounts': 0,
            'picking_policy': 'one',
            'state':'draft',		
            'user_id': uid, 			
            'order_policy': 'picking',
            'client_order_ref': "COE: %s" % trade["coe"],			
            'order_line': [],
            'note': note,			
        }
		
        post_vals = {			
            'product_uos_qty': 1,
            'product_id': 1,
            'product_uom': 1,
            'price_unit': 1,
            'product_uom_qty': 1,
            'name': '.',
            'delay': 7,
            'discount': 0,
        }		
		
        products = trade.get('products', [])
        for o in products:
            line_vals = {			
                'product_uos_qty': o.get('qty',0),
                'product_id': 1,
                'product_uom': 1,
                'price_unit': o.get('price',0),
                'product_uom_qty': o.get('qty',0),
                'name': '.',
                'delay': 7,
                'discount': 0,
            }		
			
            product_name = 	o.get('product',"").strip()		
            product_ids = self.pool.get('product.product').search(cr, uid, [('name_template','=',product_name)], context = context )
            if not product_ids:
                name = self.string_refactor(product_name)			
                product_ids = self.pool.get('product.product').search(cr, uid, [('name_template','like',name)], context = context )
                if product_ids: line_vals.update({'name': u"产品名未能精确匹配" } ) 
				
            #如果没有匹配到产品，报同步异常
            if not product_ids:
                raise osv.except_osv(u'Excel产品名错误',u'''无法再ERP中找到 - %s. \r\n请检查名字 并纠正如下情况：\r\n 1. 125 ml -> 125ml. \r\n 2. Eros - Eros Warming - 150ml  -> Eros - Warming - 150ml''' % product_name)			
                #syncerr = u"订单导入错误: 匹配不到商品。tid=%s, 商品【%s】, outer_iid=%s, num_iid=%s, outer_sku_id=%s, sku_id=%s " % ( trade.get('tid'), o.get('title', ''), o.get('outer_iid', ''), o.get('num_iid', ''),  o.get('outer_sku_id', ''), o.get('sku_id', '') )
                #self.pool.get('loewieec.error').create(cr, uid, {'name':syncerr, 'shop_id':shop.id }, context = context )
                return False
		
            #添加订单明细行			
            line_vals.update({'product_id': product_ids[0] } )
            order_val['order_line'].append( (0, 0, line_vals) ) 

        post_vals.update({'product_id': shop.gift_product_id.id , 'price_unit':0, 'product_uos_qty':1, 'product_uom_qty':1, 'product_uom': 1 } )			
        order_val['order_line'].append( (0, 0, post_vals) )		
        order_id = order_obj.create(cr, uid, order_val, context = context)

        return order_id
		
    def string_refactor(self, name):

        if not name: return False
		
        name = name.replace("/","%")
        name = name.replace("|","%")			
        ll = name.split("-")
        ll = [l.strip() for l in ll]
		
        return "%".join(ll)	
		
    def import_orders_from_excel(self, cr, uid, ids, context=None):
        attachment_obj = self.pool.get('ir.attachment')
        attachment_id = attachment_obj.search(cr,uid,[('res_id', '=', ids[0])], context=context)		
        if len(attachment_id)<1: return False

        attach = attachment_obj.browse(cr,uid,attachment_id[0],context=context)
        fname = attach.store_fname
        display_name = attach.name		
        if not fname : return False		
        fname = self.get_full_path(cr, uid, fname)
        wb = load_workbook(filename=fname)	
        #ws = wb.get_sheet_by_name("Sheet1")
        ws = wb.get_sheet_by_name(wb.get_sheet_names()[0])
        highest_row = ws.get_highest_row()
        highest_col = ws.get_highest_column()
        title_order = ws.cell(row = 1,column = 1).value	
        title_product = ws.cell(row = 1,column = 2).value
        title_qty = ws.cell(row = 1,column = 3).value
        title_coe = ws.cell(row = 1,column = 5).value		

        if highest_col < 12 or title_order != u"订单編號" or title_product != u"产品名称" or title_coe != u"COE" :
            attach.unlink()	            		
            raise osv.except_osv(u'Excel错误',u'''文件：%s 格式不正确.''' % display_name)		
			
        row_start = 2
        orders = {}	
        last_order_no = None		
        while row_start <= highest_row and ws.cell(row=row_start,column=2).value :
            line = {}
            col_start = 1    
			
            line["product"] = ws.cell(row=row_start,column=2).value
            line["qty"] = ws.cell(row=row_start,column=3).value
            cell_price = ws.cell(row=row_start,column=4)			
            line["price"] = cell_price.get_original_value()
			
            order_no = ws.cell(row=row_start,column=1).value			
            if order_no : 
                coe = ws.cell(row=row_start,column=5).value	
                name = ws.cell(row=row_start,column=6).value
                tel = ws.cell(row=row_start,column=7).value
                address = ws.cell(row=row_start,column=8).value
                city = ws.cell(row=row_start,column=9).value			
                state = ws.cell(row=row_start,column=10).value
                zip = str(ws.cell(row=row_start,column=11).value)			
                last_order_no = order_no = str(order_no)
				
                orders[last_order_no] = {"no": order_no, "coe":coe, "name":name, "tel":tel, "address":address, "city":city, "state":state, "zip":zip, "products":[line]}
            else:
                orders[last_order_no]["products"].append(line)                			
					
            row_start += 1
			
        order_ids = []		
        for key in orders.keys():		
            order_id = self.create_order_from_excel(cr, uid, ids, orders[key], context=context)	
            order_ids.append(order_id)			
        attach.unlink()		
        return order_ids		
	
    def b_sync_stock_qty(self, cr, uid, ids, context=None):

        port = 80
        shop = self.browse(cr, uid, ids[0], context = context)
        setDefaultAppInfo(shop.appkey, shop.appsecret)
        req = ProductsSearchRequest(shop.apiurl,port)
        req.fields="name, sale_props, binds, product_id, outer_id, barcode_str"
       
        #resp= req.getResponse(shop.sessionkey)
        #skus = resp.get('items_inventory_get_response',False).get('items',False).get('item', False)
        req.page_no = 1
        req.page_size = 100	
        req.q = "Fun Factory"		
        total_get = 0
        total_results = 100
        res = []
        while total_get < total_results:
            resp= req.getResponse(shop.sessionkey)
            total_results = resp.get('products_search_response').get('total_results')
			
            if total_results > 0:
                res += resp.get('products_search_response').get('products').get('product')
            total_get += req.page_size		
            req.page_no = req.page_no + 1
			
        return res		
	
    def a_sync_stock_qty(self, cr, uid, ids, context=None):

        port = 80
        shop = self.browse(cr, uid, ids[0], context = context)
        setDefaultAppInfo(shop.appkey, shop.appsecret)
        req = ItemsInventoryGetRequest(shop.apiurl,port)
        req.fields="skus, num_iid, title, type, product_id, props_name, outer_id, barcode, sub_title, shop_type, price"
       
        #resp= req.getResponse(shop.sessionkey)
        #skus = resp.get('items_inventory_get_response',False).get('items',False).get('item', False)
        req.page_no = 1
        req.page_size = 100	
        #req.is_ex = "true"
        req.is_taobao = "true"		
        total_get = 0
        total_results = 100
        res = []
        while total_get < total_results:
            resp= req.getResponse(shop.sessionkey)
            total_results = resp.get('items_inventory_get_response').get('total_results')
			
            if total_results > 0:
                res += resp.get('items_inventory_get_response').get('items').get('item')
            total_get += req.page_size		
            req.page_no = req.page_no + 1
			
        return res	
	
    def bad_sync_stock_qty(self, cr, uid, ids, context=None):
        """
         同步本条记录的库存数量到 电商店铺
        """
        port = 80
        shop = self.pool.get('loewieec.shop').browse(cr, uid, ids[0], context=context)
        for product in shop.products:
		
            qty = product.erp_stock
            if qty < 1: continue				

            setDefaultAppInfo(shop.appkey, shop.appsecret)
            req = SkusQuantityUpdateRequest(shop.apiurl, port)
            req.num_iid= long(product.ec_num_iid)

            req.skuid_quantities = "%s:%d" % (product.ec_sku_id,int(qty))
            req.type=1

            resp = req.getResponse(shop.sessionkey)

        return True	
	
    def _sync_stock_qty(self, cr, uid, ids, context=None):
        """
         同步本条记录的库存数量到 电商店铺
        """
        port = 80
        shop = self.pool.get('loewieec.shop').browse(cr, uid, ids[0], context=context)
        for product in shop.products:
		
            qty = product.erp_stock
            if qty < 1: continue				

            setDefaultAppInfo(shop.appkey, shop.appsecret)
            req = ItemQuantityUpdateRequest(shop.apiurl, port)
            req.num_iid= long(product.ec_num_iid)

            req.sku_id = product.ec_sku_id

            if qty < 0: qty = 0
            req.quantity = int(qty)
            req.type=1

            resp = req.getResponse(shop.sessionkey)

        return True
        
    def search_product_sku(self, cr, uid, ids, num_iids=None, context=None):

        port = 80
        shop = self.browse(cr, uid, ids[0], context = context)
        setDefaultAppInfo(shop.appkey, shop.appsecret)
        req = ItemSkusGetRequest(shop.apiurl,port)
        req.fields="sku_id, num_iid, properties, price, status, memo, properties_name, outer_id,quantity,barcode,change_prop,sku_spec_id,extra_id"
        
        if not num_iids : 
            req.num_iids = shop.tokenurl			
        else: 
            req.num_iids = num_iids
			
        resp= req.getResponse(shop.sessionkey)
        skus = resp.get('item_skus_get_response').get('skus', False).get('sku',False)

        return skus
	
    def search_product(self, cr, uid, ids, context=None ):	
        this = self.browse(cr, uid, ids, context=context)[0]	
        res = self._search_product(cr, uid, ids, product_name = None, start_modified=this.start_modified, end_modified=this.end_modified, context=context)

        if len(res) < 1: return 
        product_tmalljd_objs = self.pool.get('product.tmalljd') 
        num_iids = [str(o['num_iid']) for o in res]
        titles = {}
        for item in res:
            titles[item.get('num_iid')] = item.get('title') 
			
        skus_list = []		
        interval = 39
        start = end = 0
        final_end = len(num_iids) - 1
		
        while start <= final_end:
            if start + interval > final_end	: end = final_end
            else: end = start + interval			
            sub_num_iids = ",".join(num_iids[start:end])			
            skus_list +=  self.search_product_sku(cr, uid, ids, num_iids=sub_num_iids, context=context)            			
            start = end	+ 1				

        for sku in skus_list:
		
            product_vals = {'ec_shop_id':this.id,'ec_price':float(sku['price']),'ec_qty':sku['quantity'],'ec_outer_code':sku['outer_id'],'ec_sku_id':str(sku['sku_id']), 'ec_ean13':sku['barcode']}	
            num_iid = sku['num_iid']			
            product_vals['ec_num_iid'] = str(num_iid)	
            product_vals['ec_title'] = titles[num_iid]			
            pids = product_tmalljd_objs.search( cr, uid, [('ec_sku_id', '=', str(sku['sku_id'])),('ec_shop_id','=',this.id)], context=context)		
            if len(pids)>0: 
                product_tmalljd_objs.write(cr,uid,pids[0],product_vals)	            				
            else: 
                product_tmalljd_objs.create(cr, uid, product_vals)	
			
        return True	

    def _search_product(self, cr, uid, ids, product_name = None, start_modified = None, end_modified = None, context=None):
        """
        1) 按商品名称，商品修改时间搜索店铺商品
        2) start_modified、end_modified 都是UTC时间，需要加上8小时传给电商平台
        """
        shop_id = self.browse(cr, uid, ids[0], context= context)
        setDefaultAppInfo(shop_id.appkey, shop_id.appsecret)
        req = ItemsOnsaleGetRequest(shop_id.apiurl, 80)
        #req.fields="approve_status,num_iid,title,nick,outer_id,modified,num,type,price"
        req.fields="approve_status,num_iid,title,nick,type,num,list_time,price,modified,delist_time,outer_id"		
        if product_name:
            req.q = product_name
        #if start_modified:
        #    start_modified = (datetime.strptime(str(start_modified),'%Y-%m-%d %H:%M:%S',) + timedelta(hours=8)).strftime('%Y-%m-%d %H:%M:%S')
        #    req.start_modified = start_modified
        #if end_modified:
        #    end_modified = (datetime.strptime(str(end_modified),'%Y-%m-%d %H:%M:%S',) + timedelta(hours=8)).strftime('%Y-%m-%d %H:%M:%S')
        #    req.end_modified = end_modified
        
        req.page_no = 1
        req.page_size = 100
        total_get = 0
        total_results = 100
        res = []
        while total_get < total_results:
            resp= req.getResponse(shop_id.sessionkey)
            total_results = resp.get('items_onsale_get_response').get('total_results')
            #_logger.info("Jimmy total_results :%d" % total_results)			
            if total_results > 0:
                res += resp.get('items_onsale_get_response').get('items').get('item')
            total_get += req.page_size
            #_logger.info("Jimmy page_size :%d" % req.page_size)
            #_logger.info("Jimmy total_get :%d" % total_get)			
            req.page_no = req.page_no + 1
        #
        # 时间需要减去8小时
        for r in res:
            r['modified'] = (datetime.strptime(r['modified'],'%Y-%m-%d %H:%M:%S',) - timedelta(hours=8)).strftime('%Y-%m-%d %H:%M:%S')
        return res

    def create_shipping_address(self, cr, uid, partner_id, vals, context=None):
	
        partner_objs = self.pool.get('res.partner')	
        #state_objs = self.pool.get('res.country.state')
        state_id = 0		
        if 	vals.get('state', False) :	
            state_id = self.pool.get('res.country.state').search(cr,uid,[('name','like',vals.get('state'))])
		
        if not partner_id : return 0		
        partner_vals = {
            'parent_id': partner_id,		
            'name': vals.get('name'),
            'phone': vals.get('phone'),
            'street': vals.get('street'),
            'mobile': vals.get('mobile'),			
            'city': vals.get('city'),	
            'zip': vals.get('zip'),				
            'type': 'delivery',			
            'is_company': False,
            'customer': False,
            'supplier': False,
            }  
			
        if state_id : 
            partner_vals.update({'state':state_id[0]})  
			
        id = partner_objs.create(cr, uid, partner_vals, context = context )			
        return id or 0			

    def remove_duplicate_orders(self, cr, uid, orders, context=None):
        sale_obj = self.pool.get('sale.order')
        submitted_references = [o['sale_code'] for o in orders]
        existing_order_ids = sale_obj.search(cr, uid, [('name', 'in', submitted_references)], context = context)
        existing_orders = sale_obj.read(cr, uid, existing_order_ids, ['name'], context=context)
        existing_references = set([o['name'] for o in existing_orders])
        orders_to_save = [o for o in orders if o['sale_code'] not in existing_references]
        return orders_to_save
        
    def search_orders_by_modified_time(self, cr, uid, ids, status = 'WAIT_SELLER_SEND_GOODS', date_start = None, date_end = None, context=None):
        port = 80
        shop = self.browse(cr, uid, ids[0], context = context)
        partner_id = shop.partner_id.id		
        setDefaultAppInfo(shop.appkey, shop.appsecret)
        req = TradesSoldIncrementGetRequest(shop.apiurl,port)
        req.fields="type,seller_nick,buyer_nick,created,sid,tid,status,buyer_memo,seller_memo,payment,discount_fee,adjust_fee,post_fee,total_fee, pay_time,end_time,modified,received_payment,price,alipay_id,receiver_name,receiver_state,receiver_city,receiver_district,receiver_address, receiver_zip,receiver_mobile,receiver_phone,orders.price,orders.num,orders.iid,orders.num_iid,orders.sku_id,orders.refund_status,orders.status,orders.oid, orders.total_fee,orders.payment,orders.discount_fee,orders.adjust_fee,orders.sku_properties_name,orders.outer_iid,orders.outer_sku_id"
        #req.status = status
        req.type = "instant_trade,auto_delivery,guarantee_trade,tmall_i18n"		
		
        if shop.start_modified:
            #date_start = (datetime.strptime( shop.start_modified, '%Y-%m-%d %H:%M:%S',)  + timedelta(hours=8)).strftime('%Y-%m-%d %H:%M:%S')
            #req.start_modified = date_start
            req.start_modified = shop.start_modified			
        if shop.end_modified:
            #date_end = (datetime.strptime(  shop.end_modified, '%Y-%m-%d %H:%M:%S',)  + timedelta(hours=8)).strftime('%Y-%m-%d %H:%M:%S')
            #req.end_modified = date_end
            req.end_modified = shop.end_modified			
        
        res = []
        req.page_no = 1
        req.page_size = 100

        total_get = 0
        total_results = 100
        while total_get < total_results:
            resp= req.getResponse(shop.sessionkey)
            trades = resp.get('trades_sold_increment_get_response').get('trades', False)
            total_results = resp.get('trades_sold_increment_get_response').get('total_results')
            _logger.info("Jimmy total_results :%d" % total_results)				
            if total_results > 0:
                res += trades.get('trade')
            total_get += req.page_size
            req.page_no = req.page_no + 1
            _logger.info("Jimmy page_size :%d" % req.page_size)
            _logger.info("Jimmy total_get :%d" % total_get)	
        # 时间需要减去8小时
        # 单号加上店铺前缀
        order_ids = []
        for trade in res:
            trade['created'] = (datetime.strptime(trade['created'], '%Y-%m-%d %H:%M:%S',) - timedelta(hours=8)).strftime('%Y-%m-%d %H:%M:%S')
            trade['pay_time'] = (datetime.strptime(trade.get('pay_time'), '%Y-%m-%d %H:%M:%S',) - timedelta(hours=8)).strftime('%Y-%m-%d %H:%M:%S')
            trade['sale_code'] = '%s_%s' % (shop.code, trade.get('tid'))
        
        orders = self.remove_duplicate_orders(cr, uid, res, context=context)
        #orders = res		
        for trade in orders:
            #try:
            #    #_logger.info("Jimmy before create_order")		
                #order_id = self.create_order(cr, uid, ids, trade, context = context )			
            order_id = self.create_order(cr, uid, ids, trade, context = context )
            #    _logger.info("Jimmy after create_order")				
            #    order_ids.append(order_id)
            #except Exception, e:
            #    syncerr = u"店铺[%s]订单[%s]同步错误: %s" % (shop.name, trade['tid'], e)
            #    self.pool.get('loewieec.error').create(cr, uid, {'name':syncerr, 'shop_id': shop.id}, context = context )
            #    continue
        return order_ids

    def create_order(self, cr, uid, ids, trade, context=None):	
        order_obj = self.pool.get('sale.order')
        shop = self.browse(cr, uid, ids[0], context = context)	
        partner_id = shop.partner_id.id		
		
        shipping_id = self.create_shipping_address(cr, uid, partner_id, {'name':trade.get('receiver_name'),'phone':trade.get('receiver_phone'),'street':address,'mobile':trade.get('receiver_mobile')},context=context)	
        #_logger.info("Jimmy create_shipping_address : %d" % shipping_id )		
        order_val = {
            'name': "%s_%s" % (shop.code,  trade.get('tid')),
            'shop_id': shop.id,
            'date_order':  trade.get('pay_time') or datetime.now(),      #订单支付时间
            'create_date': trade.get('created'),       #订单创建时间
            'partner_id': partner_id,
            'partner_invoice_id': partner_id, 			
            'partner_shipping_id': shipping_id,
            'warehouse_id': shop.warehouse_id.id,
            'pricelist_id': shop.pricelist_id.id,
            'company_id': 1,			
            'all_discounts': 0,
            'picking_policy': 'one',
            'state':'draft',		
            'user_id': uid, 			
            'order_policy': 'picking',
            'client_order_ref': u'Loewieec_sync Generated',			
            'order_line': [],
        }
		
        post_vals = {			
            'product_uos_qty': 1,
            'product_id': 1,
            'product_uom': 1,
            'price_unit': 1,
            'product_uom_qty': 1,
            'name': '.',
            'delay': 7,
            'discount': 0,
        }		
		
        orders = trade.get('orders', {}).get('order', [])
        for o in orders:
            line_vals = {			
                'product_uos_qty': 1,
                'product_id': 1,
                'product_uom': 1,
                'price_unit': 1,
                'product_uom_qty': 1,
                'name': '.',
                'delay': 7,
                'discount': 0,
            }			
            product_tmalljd_ids = self.pool.get('product.tmalljd').search(cr, uid, [('ec_sku_id','=',o.get('sku_id'))], context = context )
            #_logger.info("Jimmy num_iid : %s" % o.get('num_iid') )            
            #如果没有匹配到产品，报同步异常
            if not product_tmalljd_ids:
                syncerr = u"订单导入错误: 匹配不到商品。tid=%s, 商品【%s】, outer_iid=%s, num_iid=%s, outer_sku_id=%s, sku_id=%s " % ( trade.get('tid'), o.get('title', ''), o.get('outer_iid', ''), o.get('num_iid', ''),  o.get('outer_sku_id', ''), o.get('sku_id', '') )
                self.pool.get('loewieec.error').create(cr, uid, {'name':syncerr, 'shop_id':shop.id }, context = context )
                return False

            product_tmalljd_obj = self.pool.get('product.tmalljd').browse(cr, uid, product_tmalljd_ids[0], context = context)
            product_id = product_tmalljd_obj.erp_product_id.id
            uom_id = product_tmalljd_obj.erp_product_id.uom_id.id			
            #添加订单明细行
            price_unit = float(o.get('payment',0))/o.get('num') or float(o.get('total_fee',0))/o.get('num')			
            line_vals.update({'product_id': product_id , 'price_unit':price_unit, 'product_uos_qty':o.get('num'), 'product_uom_qty':o.get('num'), 'product_uom': uom_id } )
            order_val['order_line'].append( (0, 0, line_vals) ) 

        post_vals.update({'product_id': shop.gift_product_id.id , 'price_unit':0, 'product_uos_qty':1, 'product_uom_qty':1, 'product_uom': 1 } )			
        order_val['order_line'].append( (0, 0, post_vals) )		
        order_id = order_obj.create(cr, uid, order_val, context = context)

        return order_id
		
    def get_tmall_time(self, cr, uid, ids, context=None ):	
        shop = self.browse(cr, uid, ids[0], context = context)	
        setDefaultAppInfo(shop.appkey, shop.appsecret)
        req = TimeGetRequest(shop.apiurl,80)	
        resp= req.getResponse(shop.sessionkey)
        ttime = resp.get('time_get_response').get('time',False)
		
        if ttime :
            shop.tmall_time = ttime
        return True			
		
    def search_orders_by_created_time(self, cr, uid, ids, status = 'WAIT_SELLER_SEND_GOODS', date_start = None, date_end = None, context=None):
        port = 80
        shop = self.browse(cr, uid, ids[0], context = context)
        partner_id = shop.partner_id.id		
        setDefaultAppInfo(shop.appkey, shop.appsecret)
        req = TradesSoldGetRequest(shop.apiurl,port)
        req.fields="type,seller_nick,buyer_nick,created,sid,tid,status,buyer_memo,seller_memo,payment,discount_fee,adjust_fee,post_fee,total_fee, pay_time,end_time,modified,received_payment,price,alipay_id,receiver_name,receiver_state,receiver_city,receiver_district,receiver_address, receiver_zip,receiver_mobile,receiver_phone,orders.price,orders.num,orders.iid,orders.num_iid,orders.sku_id,orders.refund_status,orders.status,orders.oid, orders.total_fee,orders.payment,orders.discount_fee,orders.adjust_fee,orders.sku_properties_name,orders.outer_iid,orders.outer_sku_id"
        #req.status = status
        req.type = "instant_trade,auto_delivery,guarantee_trade,tmall_i18n"		
		
        #if shop.start_modified:
            #date_start = (datetime.strptime( shop.start_modified, '%Y-%m-%d %H:%M:%S',)  + timedelta(hours=8)).strftime('%Y-%m-%d %H:%M:%S')
            #req.start_modified = date_start
            #req.start_created = shop.start_modified	
        #req.start_created = (datetime.now()- timedelta(hours=64)).strftime('%Y-%m-%d %H:%M:%S')	
        req.start_created = datetime.now().strftime('%Y-%m-%d %H:%M:%S')		
        #if shop.end_modified:
            #date_end = (datetime.strptime(  shop.end_modified, '%Y-%m-%d %H:%M:%S',)  + timedelta(hours=8)).strftime('%Y-%m-%d %H:%M:%S')
            #req.end_modified = date_end
            #req.end_created = shop.end_modified			
        req.end_created = (datetime.now()+ timedelta(hours=8)).strftime('%Y-%m-%d %H:%M:%S')   
		
        res = []
        req.page_no = 1
        req.page_size = 100

        total_get = 0
        total_results = 100
        while total_get < total_results:
            resp= req.getResponse(shop.sessionkey)
            trades = resp.get('trades_sold_get_response').get('trades', False)
            total_results = resp.get('trades_sold_get_response').get('total_results')
            _logger.info("Jimmy total_results :%d" % total_results)				
            if total_results > 0:
                res += trades.get('trade')
            total_get += req.page_size
            req.page_no = req.page_no + 1
            _logger.info("Jimmy page_size :%d" % req.page_size)
            _logger.info("Jimmy total_get :%d" % total_get)	
        # 时间需要减去8小时
        # 单号加上店铺前缀
        order_ids = []
        for trade in res:
            trade['created'] = (datetime.strptime(trade.get('created'), '%Y-%m-%d %H:%M:%S',) - timedelta(hours=8)).strftime('%Y-%m-%d %H:%M:%S')
            trade['pay_time'] = (datetime.strptime((trade.get('pay_time',False) or '2016-01-01 00:00:01'), '%Y-%m-%d %H:%M:%S',) - timedelta(hours=8)).strftime('%Y-%m-%d %H:%M:%S')
            trade['sale_code'] = '%s_%s' % (shop.code, trade.get('tid'))
        
        orders = self.remove_duplicate_orders(cr, uid, res, context=context)
        #orders = res		
        for trade in orders:
            #try:
            #    #_logger.info("Jimmy before create_order")		
                #order_id = self.create_order(cr, uid, ids, trade, context = context )			
            order_id = self.create_order(cr, uid, ids, trade, context = context )
            #    _logger.info("Jimmy after create_order")				
            #    order_ids.append(order_id)
            #except Exception, e:
            #    syncerr = u"店铺[%s]订单[%s]同步错误: %s" % (shop.name, trade['tid'], e)
            #    self.pool.get('loewieec.error').create(cr, uid, {'name':syncerr, 'shop_id': shop.id}, context = context )
            #    continue
        return order_ids
		
    def search_orders_by_tid(self, cr, uid, ids, context=None):
        port = 80
        shop = self.browse(cr, uid, ids[0], context = context)
        partner_id = shop.partner_id.id		
        setDefaultAppInfo(shop.appkey, shop.appsecret)
        req = TradeFullinfoGetRequest(shop.apiurl,port)
        #req.fields="type,seller_nick,buyer_nick,created,sid,tid,status,buyer_memo,seller_memo,payment,discount_fee,adjust_fee,post_fee,total_fee, pay_time,end_time,modified,received_payment,price,alipay_id,receiver_name,receiver_state,receiver_city,receiver_district,receiver_address, receiver_zip,receiver_mobile,receiver_phone,orders.price,orders.num,orders.iid,orders.num_iid,orders.sku_id,orders.refund_status,orders.status,orders.oid, orders.total_fee,orders.payment,orders.discount_fee,orders.adjust_fee,orders.sku_properties_name,orders.outer_iid,orders.outer_sku_id"
        req.fields = "type,receiver_name,receiver_state,receiver_city,receiver_district,receiver_address, receiver_zip,receiver_mobile,receiver_phone"
        req.tid = shop.authurl	

        resp= req.getResponse(shop.sessionkey)
        trade = resp.get('trade_fullinfo_get_response').get('trade', False)
        _logger.info("Jimmy type :%s" % trade.get('type'))
        _logger.info("Jimmy receiver_name :%s" % trade.get('receiver_name'))
        _logger.info("Jimmy receiver_address :%s" % trade.get('receiver_address'))
        _logger.info("Jimmy receiver_mobile :%s" % trade.get('receiver_mobile'))
        _logger.info("Jimmy receiver_phone :%s" % trade.get('receiver_phone'))		
               			
        return True
		
		
				
		
		