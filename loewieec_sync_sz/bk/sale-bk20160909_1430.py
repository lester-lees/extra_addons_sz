# -*- encoding: utf-8 -*-
from openerp.osv import fields,osv
from openpyxl.reader.excel import load_workbook 
from openerp import tools
import os
import re

class product_tmalljd(osv.osv):
    _name = "product.tmalljd"
    #_inherits = {'product.product': 'product_id'}

    def _get_ean13(self, cr, uid, ids, field_name, arg, context=None):	
        result = {}
        for line in self.pool.get('product.tmalljd').browse(cr, uid, ids, context=context):
            if line.erp_product_id : result[line.id] = line.erp_product_id.ean13 or line.erp_product_id.default_code
        return result		

    def _get_stock(self, cr, uid, ids, field_name, arg, context=None):	
        result = {}
        domain_products = [('location_id','=',38)]		
        quants = self.pool.get('stock.quant').read_group(cr, uid, domain_products, ['product_id', 'qty'], ['product_id'], context=context)	
        quants = dict(map(lambda x: (x['product_id'][0], x['qty']), quants))
		
        for line in self.pool.get('product.tmalljd').browse(cr, uid, ids, context=context):
            id = line.id
            if line.erp_product_id :
                pid = line.erp_product_id.id			
                result[id] = quants.get(pid, 0.0)
            else:
                result[id] = 0			
        return result		
		
    _columns = {
        'erp_product_id': fields.many2one('product.product','ERP Product Name'),
        'erp_ean13': fields.char('ERP_EAN13'), #fields.function(_get_ean13,type='char',string='ERP_EAN13'),
        'erp_stock': fields.float('ERPStock'),#fields.function(_get_stock,type='float',string='ERP库存'),		
        'ec_shop_id': fields.many2one('loewieec.shop', u'店铺'),		
        'ec_num_iid': fields.char(u'淘宝数字编码'),
        'ec_sku_id': fields.char(u'淘宝SKU_ID'),
        'ec_title':fields.char(u'商品标题'),
        'ec_price':fields.float(u'售价'),
        'ec_color':fields.char(u'颜色'),
        'ec_ean13': fields.char(u'条形码'),
        'ec_brand': fields.char(u'品牌'),
        'ec_qty': fields.integer(u'EC数量'),
        'ec_outer_code': fields.char(u'商家外部编码'),		
        'ec_product_name': fields.char(u'产品名称'),	
        'ec_product_id': fields.char(u'EC产品ID'),		
        'ec_num_custom':fields.char(u'海关代码'),	
    }
	
class loewieec_error(osv.osv):
    _name = "loewieec.error"

    _columns = {	
        'shop_id': fields.many2one('loewieec.shop', u'店铺'),
        'name': fields.char(u'错误信息'),
    }
	
class sale_coe(osv.osv):
    _name = "sale.coe"	
	
    _columns = {	
        'sale_id':fields.many2one('sale.order',string='Sales Order'),	
        'tmi_jdi_no': fields.char(string=u'电商单号'),		
        'express_no':fields.char(string=u'货运单号'),	
        'expresser':fields.char(string=u'货运公司'),		
        'name': fields.char(string=u'收货人'),	
        'receive_name': fields.char(string=u'收货人old'),		
        'tel': fields.char(string=u'电话'),
        'mobile': fields.char(string=u'手机'),		
        'state': fields.char(string=u'省份'),	
        'city': fields.char(string=u'城市'),	
        'address': fields.char(string=u'地址'),	
        'zip': fields.char(string=u'邮编'),	
        'weight': fields.float(string=u'重量'),		
        'price': fields.float(string=u'快递费'),		
    }
	
class sale_order_line(osv.osv):
    _inherit = "sale.order.line"
	
    _columns = {
        'express_id': fields.many2one('sale.coe',string=u'快递信息'),	
        'tmi_jdi_no': fields.char(string=u'电商单号'),
        'pay_time': fields.datetime('PayTime'),	
        'create_time_tmjd': fields.datetime('TM_JD Create Time'),		
    }	

class stock_move(osv.osv):
    _inherit = "stock.move"
	
    def _get_express_id(self, cr, uid, ids, field_name, arg, context=None):	
        result = {}
        for move in self.pool.get('stock.move').browse(cr, uid, ids, context=context):
            result[move.id] = move.procurement_id.sale_line_id.express_id.id
        return result	
		
    _columns = {
        #'sale_order_line': fields.function(_get_sale_order_line, type='char',string='Sales Line'),	
        'express_id': fields.function(_get_express_id,type='many2one',relation='sale.coe',string='COE NO'),
    }	
	
class sale_order(osv.osv):
    _name = "sale.order"
    _inherit = "sale.order"
        
    _columns = {
        #'express_ids': fields.function(_get_express_ids,type='one2many',string=""),
        'note2':fields.html('Note2'),		
        'selected': fields.boolean('Selected'),	
        'shop_id': fields.many2one('loewieec.shop', string=u"EC店铺名", readonly=True),
        'sale_code': fields.char(u'EC单号', readonly=True),				
        'order_state': fields.selection([
            ('WAIT_SELLER_SEND_GOODS', u'等待卖家发货'),
            ('WAIT_BUYER_CONFIRM_GOODS', u'等待买家确认收货'),
            ('TRADE_FINISHED', u'交易成功'),
            ('TRADE_CLOSED', u'交易关闭'),
            ], u'订单状态'),
    }
        	
    def get_full_path(self, cr, uid, path):
        path = re.sub('[.]', '', path)
        path = path.strip('/\\')
        return os.path.join(tools.config.filestore(cr.dbname), path)
		
    def string_refactor(self, name):
        if not name: return False		
        name = name.replace("/","%")
        name = name.replace("|","%")			
        ll = name.split("-")
        ll = [l.strip() for l in ll]
        ll = "%".join(ll)				
        return ll.replace(" ","%")
		
    def import_orders_from_excel(self, cr, uid, ids, context=None):
        sale_order_obj = self.pool.get('sale.order').browse(cr,uid,ids[0],context=context)
        note = sale_order_obj.note		
        lines = note.split(chr(10))
        for line in lines:
            ll = line.split(chr(9))		

        return		
		
    def import_orders_from_excel_bk(self, cr, uid, ids, context=None):	
        attachment_obj = self.pool.get('ir.attachment')
        attachment_id = attachment_obj.search(cr,uid,[('res_id', '=', ids[0])], context=context)		
        if len(attachment_id)<1: return False

        attach = attachment_obj.browse(cr,uid,attachment_id[0],context=context)
        fname = attach.store_fname
        display_name = attach.name		
        if not fname : return False		
        fname = self.get_full_path(cr, uid, fname)
        wb = load_workbook(filename=fname)	
        ws = wb.get_sheet_by_name('order')  or wb.get_sheet_by_name( wb.get_sheet_names()[0] ) 
		
        highest_row = ws.get_highest_row()
        highest_col = ws.get_highest_column()
        product_name = ws.cell(row = 0,column = 1).value	
        product_qty = ws.cell(row = 0,column = 2).value
        receive_name = ws.cell(row = 0,column = 3).value
        cellphone = ws.cell(row = 0,column = 4).value		

        if highest_col < 9 or product_name != u"产品名称 - 此列可自动提示输入产品名" or product_qty != u"数量" or receive_name != u"姓名"  or cellphone != u"電話" :	            		
            raise osv.except_osv(u'Excel错误',u'''文件：%s 格式不正确.''' % display_name)		
			
        row_start = 1
        recipients = {}	
        last_mobile = None		
        while row_start <= highest_row and ws.cell(row=row_start,column=1).value :	
            name = ws.cell(row=row_start,column=3).value		
            mobile = str( ws.cell(row=row_start,column=4).value )
            address = ws.cell(row=row_start,column=5).value	
            city = ws.cell(row=row_start,column=6).value
            state = ws.cell(row=row_start,column=7).value
            zip = ws.cell(row=row_start,column=8).value			
			
            if bool(name) != bool(mobile) :            		
                raise osv.except_osv(u'Excel错误',u'''文件'%s'第：%d行不正确地合并单元格.''' % (display_name, row_start+1) )  
				
            if mobile not in recipients.keys(): 		
                last_mobile = mobile				
                recipients[mobile] = {"id":0, "sale_id":1, "name": name, "mobile":mobile, "address":address, "city":city, "state":state, "zip":zip}
				
            row_start += 1

        row_start = 1
        product_lines = []
        last_mobile = None		
        while row_start <= highest_row and ws.cell(row=row_start,column=2).value :
            line = {}
            mobile = str( ws.cell(row=row_start,column=4).value )
            line['mobile'] = mobile			
            line["product_id"] = 1			
            line["product"] = ws.cell(row=row_start,column=1).value
            line["desc"] = "-"	
            qty_tmp = ws.cell(row=row_start,column=2)			
            line["qty"] = qty_tmp.get_original_value() or 1	
			
            product_lines.append(line)					
            row_start += 1				
	
        self.create_sale_order_line(cr, uid, ids, recipients, product_lines, display_name, context=context)			
        attach.unlink()		
        return True		


    def create_sale_order_line(self, cr, uid, ids, recipients, product_lines, tmijdi_fname, context=None):	
        order_obj = self.pool.get('sale.order').browse(cr,uid,ids[0],context=context)
        if not order_obj : 
            raise osv.except_osv(u'Sale order 错误',u'''请先保存销售单草稿''')	
            return False
		
        product_obj = self.pool.get('product.product')		
        for product in product_lines:			
            product_name = 	product.get('product',"").strip()		
            product_ids = product_obj.search(cr, uid, [('name_template','=ilike',product_name)], context = context )
            if not product_ids:
                name = self.string_refactor(product_name)			
                product_ids = product_obj.search(cr, uid, [('name_template','ilike',name)], context = context )
                if product_ids: product['desc'] = u"产品名未能精确匹配" 
				
            if not product_ids:
                raise osv.except_osv(u'Excel产品名错误',u'''无法再ERP中找到 - %s. \r\n请检查名字 并纠正如下情况：\r\n 1. 125 ml -> 125ml. \r\n 2. Eros - Eros Warming - 150ml  -> Eros - Warming - 150ml''' % product_name)			
                return False
            product['product_id'] = product_ids[0]	
			
        note = ""			
        coe_obj = self.pool.get('sale.coe')			
        for key in recipients:
            coe_line = recipients[key]		
            coe_vals = {	# 为每个收件人 创建COE 条目 		
                'sale_id': ids[0],
                'name': coe_line["name"],
                'mobile': coe_line["mobile"],
                'state': coe_line["state"],
                'city': coe_line["city"],
                'address': coe_line["address"],
                'zip': coe_line["zip"],
            }
            coe_id = coe_obj.search(cr, uid, [('name','=',coe_line["name"]),('mobile','=',coe_line["mobile"])], context=context)
            coe_id = coe_id and coe_id[0] or 0			
            if not coe_id :			
                coe_id = coe_obj.create(cr, uid, coe_vals, context=context)	
            coe_line["id"] = coe_id	           	
		
            note += u"姓名:%s, 电话:%s, 省份:%s, 城市%s, 收货地址:%s, 邮编:%s;"	% (coe_line["name"],coe_line["mobile"],coe_line["state"],coe_line["city"],coe_line["address"],coe_line["zip"] )
            note += chr(10)	  
			
        sale_order_line_obj = self.pool.get('sale.order.line')	
        pricelist_id = order_obj.pricelist_id and order_obj.pricelist_id.id or 0		
        for o in product_lines:	
            price = 0		
            if pricelist_id :		
                price = sale_order_line_obj.get_price(cr, pricelist_id=pricelist_id, product_id=o.get('product_id'))		
            line_vals = {
                'sale_id': ids[0],			
                'product_uos_qty': o.get('qty',1),
                'product_id': o.get('product_id'),
                'product_uom': 1,				
                'price_unit': price,
                'product_uom_qty': o.get('qty',1),
                'name': o.get('desc',0),
                'delay': 7,
                'discount': 0,
                'express_id':recipients[o.get('mobile')]['id'],				
            }				
            sale_order_line_obj.create(cr, uid, line_vals, context = context)

        return True


	