# -*- encoding: utf-8 -*-
from openerp.osv import fields,osv
from openpyxl.reader.excel import load_workbook 
from openerp import tools
from openerp.tools.translate import _
import os
import re

class product_tmalljd(osv.osv):
    _name = "product.tmalljd"
    #_inherits = {'product.product': 'product_id'}

    def _get_ean13(self, cr, uid, ids, field_name, arg, context=None):	
        result = {}
        for line in self.pool.get('product.tmalljd').browse(cr, uid, ids, context=context):
            if line.erp_product_id : result[line.id] = line.erp_product_id.ean13 or line.erp_product_id.default_code
        return result		

    def _get_stock(self, cr, uid, ids, field_name, arg, context=None):	
        result = {}
        domain_products = [('location_id','=',38)]		
        quants = self.pool.get('stock.quant').read_group(cr, uid, domain_products, ['product_id', 'qty'], ['product_id'], context=context)	
        quants = dict(map(lambda x: (x['product_id'][0], x['qty']), quants))
		
        for line in self.pool.get('product.tmalljd').browse(cr, uid, ids, context=context):
            id = line.id
            if line.erp_product_id :
                pid = line.erp_product_id.id			
                result[id] = quants.get(pid, 0.0)
            else:
                result[id] = 0			
        return result		
		
    _columns = {
        'erp_product_id': fields.many2one('product.product','ERP Product Name'),
        'erp_ean13': fields.char('ERP_EAN13'), #fields.function(_get_ean13,type='char',string='ERP_EAN13'),
        'erp_stock': fields.float('ERPStock'),#fields.function(_get_stock,type='float',string='ERP库存'),		
        'ec_shop_id': fields.many2one('loewieec.shop', u'店铺'),		
        'ec_num_iid': fields.char(u'淘宝数字编码'),
        'ec_sku_id': fields.char(u'淘宝SKU_ID'),
        'ec_title':fields.char(u'商品标题'),
        'ec_price':fields.float(u'售价'),
        'ec_color':fields.char(u'颜色'),
        'ec_ean13': fields.char(u'条形码'),
        'ec_brand': fields.char(u'品牌'),
        'ec_qty': fields.integer(u'EC数量'),
        'ec_outer_code': fields.char(u'商家外部编码'),		
        'ec_product_name': fields.char(u'产品名称'),	
        'ec_product_id': fields.char(u'EC产品ID'),		
        'ec_num_custom':fields.char(u'海关代码'),	
    }
	
class loewieec_error(osv.osv):
    _name = "loewieec.error"

    _columns = {	
        'shop_id': fields.many2one('loewieec.shop', u'店铺'),
        'name': fields.char(u'错误信息'),
    }
	
class sale_coe(osv.osv):
    _name = "sale.coe"	
	
    _columns = {	
        'sale_id':fields.many2one('sale.order',string='Sales Order'),	
        'tmi_jdi_no': fields.char(string=u'电商单号'),		
        'express_no':fields.char(string=u'货运单号'),	
        'expresser':fields.char(string=u'货运公司'),		
        'name': fields.char(string=u'收货人'),	
        'receive_name': fields.char(string=u'收货人old'),		
        'tel': fields.char(string=u'电话'),
        'mobile': fields.char(string=u'手机'),		
        'state': fields.char(string=u'省份'),	
        'city': fields.char(string=u'城市'),	
        'address': fields.char(string=u'地址'),	
        'zip': fields.char(string=u'邮编'),	
        'weight': fields.float(string=u'重量'),		
        'price': fields.float(string=u'快递费'),		
    }
	
class sale_order_line(osv.osv):
    _inherit = "sale.order.line"
	
    _columns = {
        'express_id': fields.many2one('sale.coe',string=u'快递信息'),	
        'tmi_jdi_no': fields.char(string=u'电商单号'),
        'pay_time': fields.datetime('PayTime'),	
        'create_time_tmjd': fields.datetime('TM_JD Create Time'),		
    }	

class stock_move(osv.osv):
    _inherit = "stock.move"
	
    def _get_express_id(self, cr, uid, ids, field_name, arg, context=None):	
        result = {}
        for move in self.pool.get('stock.move').browse(cr, uid, ids, context=context):
            result[move.id] = move.procurement_id.sale_line_id.express_id.id
        return result	
		
    _columns = {
        #'sale_order_line': fields.function(_get_sale_order_line, type='char',string='Sales Line'),	
        'express_id': fields.function(_get_express_id,type='many2one',relation='sale.coe',string='COE NO'),
    }	

class stock_picking(osv.osv):
    _inherit = "stock.picking"
	
    def view_express_data(self, cr, uid, ids, context=None):	
        stock_picking_obj = self.pool.get('stock.picking').browse(cr,uid,ids[0],context=context)	
        if not stock_picking_obj.sale_id : 
            raise osv.except_osv(u'stock.picking 错误',u'''没有销售单与此仓库单有关联''')	
            return False
        order_id = 	stock_picking_obj.sale_id.id		
        sale_coe_obj = self.pool.get('sale.coe')
        express_ids = sale_coe_obj.search(cr,uid,[('sale_id','=',order_id)],context=context)		
        if len(express_ids) < 1: return False
		
        view_ref = self.pool.get('ir.model.data').get_object_reference(cr, uid, 'loewieec_sync', 'ecshop_sale_coe_tree')
        view_id = view_ref and view_ref[1] or False
        if not view_id : return False		
        return {
            'name': _('Express Info'),
            'type': 'ir.actions.act_window',
            'view_type': 'tree',
            'view_mode': 'tree',
            'res_model': 'sale.coe',
            'view_id': view_id,
            'target': 'current',
            'res_id': express_ids,
            'domain': [('sale_id','=',order_id)],					
        }        		
	
class sale_order(osv.osv):
    _name = "sale.order"
    _inherit = "sale.order"
        
    _columns = {
        #'express_ids': fields.function(_get_express_ids,type='one2many',string=""),
        #'note2':fields.html('Note2'),		
        'selected': fields.boolean('Selected'),	
        'shop_id': fields.many2one('loewieec.shop', string=u"EC店铺名", readonly=True),
        'sale_code': fields.char(u'EC单号', readonly=True),				
        'order_state': fields.selection([
            ('WAIT_SELLER_SEND_GOODS', u'等待卖家发货'),
            ('WAIT_BUYER_CONFIRM_GOODS', u'等待买家确认收货'),
            ('TRADE_FINISHED', u'交易成功'),
            ('TRADE_CLOSED', u'交易关闭'),
            ], u'订单状态'),
    }
        	
    def get_full_path(self, cr, uid, path):
        path = re.sub('[.]', '', path)
        path = path.strip('/\\')
        return os.path.join(tools.config.filestore(cr.dbname), path)
		
    def string_refactor(self, name):
        if not name: return False		
        name = name.replace("/","%")
        name = name.replace("|","%")			
        ll = name.split("-")
        ll = [l.strip() for l in ll]
        ll = "%".join(ll)				
        return ll.replace(" ","%")
				
		
    def import_orders_from_excel_bk(self, cr, uid, ids, context=None):	
        attachment_obj = self.pool.get('ir.attachment')
        attachment_id = attachment_obj.search(cr,uid,[('res_id', '=', ids[0])], context=context)		
        if len(attachment_id)<1: return False

        attach = attachment_obj.browse(cr,uid,attachment_id[0],context=context)
        fname = attach.store_fname
        display_name = attach.name		
        if not fname : return False		
        fname = self.get_full_path(cr, uid, fname)
        wb = load_workbook(filename=fname)	
        ws = wb.get_sheet_by_name('order')  or wb.get_sheet_by_name( wb.get_sheet_names()[0] ) 
		
        highest_row = ws.get_highest_row()
        highest_col = ws.get_highest_column()
        product_name = ws.cell(row = 0,column = 1).value	
        product_qty = ws.cell(row = 0,column = 2).value
        receive_name = ws.cell(row = 0,column = 3).value
        cellphone = ws.cell(row = 0,column = 4).value		

        if highest_col < 9 or product_name != u"产品名称 - 此列可自动提示输入产品名" or product_qty != u"数量" or receive_name != u"姓名"  or cellphone != u"電話" :	            		
            raise osv.except_osv(u'Excel错误',u'''文件：%s 格式不正确.''' % display_name)		
			
        row_start = 1
        recipients = {}	
        last_mobile = None		
        while row_start <= highest_row and ws.cell(row=row_start,column=1).value :	
            name = ws.cell(row=row_start,column=3).value		
            mobile = str( ws.cell(row=row_start,column=4).value )
            address = ws.cell(row=row_start,column=5).value	
            city = ws.cell(row=row_start,column=6).value
            state = ws.cell(row=row_start,column=7).value
            zip = ws.cell(row=row_start,column=8).value			
			
            if bool(name) != bool(mobile) :            		
                raise osv.except_osv(u'Excel错误',u'''文件'%s'第：%d行不正确地合并单元格.''' % (display_name, row_start+1) )  
				
            if mobile not in recipients.keys(): 		
                last_mobile = mobile				
                recipients[mobile] = {"id":0, "sale_id":1, "name": name, "mobile":mobile, "address":address, "city":city, "state":state, "zip":zip}
				
            row_start += 1

        row_start = 1
        product_lines = []
        last_mobile = None		
        while row_start <= highest_row and ws.cell(row=row_start,column=2).value :
            line = {}
            mobile = str( ws.cell(row=row_start,column=4).value )
            line['mobile'] = mobile			
            line["product_id"] = 1			
            line["product"] = ws.cell(row=row_start,column=1).value
            line["desc"] = "-"	
            qty_tmp = ws.cell(row=row_start,column=2)			
            line["qty"] = qty_tmp.get_original_value() or 1	
			
            product_lines.append(line)					
            row_start += 1				
	
        self.create_sale_order_line(cr, uid, ids, recipients, product_lines, display_name, context=context)			
        attach.unlink()		
        return True		

		
    def view_express_data(self, cr, uid, ids, context=None):	
        sale_order_obj = self.pool.get('sale.order').browse(cr,uid,ids[0],context=context)	
        if not sale_order_obj : 
            raise osv.except_osv(u'Sale order 错误',u'''请先保存销售单草稿''')	
            return False
        sale_coe_obj = self.pool.get('sale.coe')
        express_ids = sale_coe_obj.search(cr,uid,[('sale_id','=',ids[0])],context=context)		
        if len(express_ids) < 1: return False

        mod_obj = self.pool.get('ir.model.data')
        act_obj = self.pool.get('ir.actions.act_window')

        result = mod_obj.get_object_reference(cr, uid, 'loewieec_sync', 'action_loewieec_salecoe')
        id = result and result[1] or False
        result = act_obj.read(cr, uid, [id], context=context)[0]
        result['domain'] = [('sale_id','=',ids[0])]
        result['res_id'] = express_ids
		
        return result		
		
    def import_orders_from_note(self, cr, uid, ids, context=None):	
        sale_order_obj = self.pool.get('sale.order').browse(cr,uid,ids[0],context=context)	
        if not sale_order_obj : 
            raise osv.except_osv(u'Sale order 错误',u'''请先保存销售单草稿''')	
            return False
	
        lines = [ o.split(chr(9)) for o in sale_order_obj.note.split(chr(10))]				
        if lines[0][0] == u'序号' and lines[0][5] == u'收件人地址' : 
            lines.remove(lines[0])
        note = ""			
        coe_obj = self.pool.get('sale.coe')	
        express = {}		
        for line in lines:
            if len(line)<8: 
                lines.remove(line)		
                continue				
            if line[4] in express.keys(): continue
			
            vals = {	# 为每个收件人 创建COE 条目 		
                'sale_id': ids[0],
                'id':0,				
                'name':line[3], #str(line[3]).strip(),
                'mobile':line[4], #str(line[4]).strip(),
                'address':line[5], #str(line[5]).strip(),	
                'city':line[6], #str(line[6]).strip(),
                'state':line[7], #str(line[7]).strip(),				
                'zip':line[8], #str(line[8]).strip(),
            }
            coe_id = coe_obj.search(cr, uid, [('name','=',vals["name"]),('mobile','=',vals['mobile']),('sale_id','=',ids[0])], context=context)
            coe_id = coe_id and coe_id[0] or 0			
            if not coe_id :			
                coe_id = coe_obj.create(cr, uid, vals, context=context)	
            vals["id"] = coe_id	           	
		
            note += u"姓名:%s, 电话:%s, 省份:%s, 城市%s, 收货地址:%s, 邮编:%s;"	% (vals["name"],vals["mobile"],vals["state"],vals["city"],vals["address"],vals["zip"] )
            note += chr(10)	 
            express[vals["mobile"]]	= vals		
			
        product_obj = self.pool.get('product.product')				
        sale_order_line_obj = self.pool.get('sale.order.line')			
        for product in lines:	
            product_name = 	product[1].strip()		
            product_ids = product_obj.search(cr, uid, [('name_template','=ilike',product_name)], context = context )
            if not product_ids:
                name = self.string_refactor(product_name)			
                product_ids = product_obj.search(cr, uid, [('name_template','ilike',name)], context = context )
                if product_ids: product[3] = u"产品名未能精确匹配" 
				
            if not product_ids:
                raise osv.except_osv(u'Excel产品名错误',u'''无法再ERP中找到 - %s. \r\n请检查名字 并纠正如下情况：\r\n 1. 125 ml -> 125ml. \r\n 2. Eros - Eros Warming - 150ml  -> Eros - Warming - 150ml''' % product_name)			
                return False
            product[0] = product_ids[0]	

        pricelist_id = sale_order_obj.pricelist_id and sale_order_obj.pricelist_id.id or 0			
        for o in lines:				
            price = 0		
            if pricelist_id :		
                price = sale_order_line_obj.get_price(cr, pricelist_id=pricelist_id, product_id=o[0])		
            line_vals = {
                'order_id': ids[0],			
                'product_uos_qty': int(o[2]),
                'product_id': o[0],
                'product_uom': 1,				
                'price_unit': price,
                'product_uom_qty': int(o[2]),
                'name': o[3],
                'delay': 7,
                'discount': 0,
                'express_id':express[o[4]]['id'],				
            }				
            sale_order_line_obj.create(cr, uid, line_vals)

        return True


	