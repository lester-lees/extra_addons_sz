# -*- coding: utf-8 -*-

from openerp.osv import fields, osv
from openerp import tools
import os
import re
from openpyxl.reader.excel import load_workbook 
import logging
_logger = logging.getLogger(__name__)

class stock_inventory_line(osv.osv):
    _inherit = "stock.inventory.line"

    #def _qty_ec(self, cr, uid, ids, field_names=None, arg=False, context=None):	
	
    _columns = {
        'total_all': fields.integer(string=u'盘点总计(电商+成品)'),	
        'qty_ec': fields.related('product_id', 'qty_onhand_ec', relation='product.product', type='integer',string=u"电商在手"),		
    }

	
class stock_inventory(osv.osv):
    _inherit = "stock.inventory"
    _order = "id desc,date desc, name"	
    _defaults = {
        'filter': 'product',
        'product_id':19374,		
    }
	
    def get_full_path(self, cr, uid, path):
        # sanitize ath
        path = re.sub('[.]', '', path)
        path = path.strip('/\\')
        return os.path.join(tools.config.filestore(cr.dbname), path)
		
    def import_from_excel(self, cr, uid, ids, context=None):
        attachment_obj = self.pool.get('ir.attachment')
        attachment_id = attachment_obj.search(cr,uid,[('res_id', '=', ids[0])], context=context)		
        if len(attachment_id)<1: return False
        attach = attachment_obj.browse(cr,uid,attachment_id[0],context=context)
        fname = attach.store_fname
        display_name = attach.name		
        if not fname : return False		
        fname = self.get_full_path(cr, uid, fname)

        wb = load_workbook(filename=fname)        
        ws = wb.get_sheet_by_name(wb.get_sheet_names()[0])
        highest_row = ws.get_highest_row()
        highest_col = ws.get_highest_column()
        title_index = ws.cell(row = 0,column = 0).value	
        title_brand = ws.cell(row = 0,column = 1).value		
        title_real_qty = ws.cell(row = 0,column = 5).value

        if highest_col < 6 or title_index != u"品牌" or title_brand != u"产品名" or title_real_qty != u"盘点数" :	            		
            raise osv.except_osv(u'Excel错误',u'''格式错误.''' % display_name)
			
        product_obj = self.pool.get('product.product')	
        row_start = 1		
        error_list = []	
        lines_list = []		
        while row_start <= highest_row and ws.cell(row=row_start,column=1).value :	
			
            name = ws.cell(row=row_start,column=1).value
			
            if not name	: 
                error_list.append('Line:%d, product name is empty.' % row_start )
                row_start += 1
                continue
				
            name = name.strip()			
            qty_tmp = ws.cell(row=row_start,column=5)			
            qty = qty_tmp.get_original_value() or 0	
			
            pid = product_obj.search(cr,uid,[('name_template','=',name)],context=context)
            if not pid: 			
                error_list.append('Line:%d, no matching ERP Product Name. ' % row_start )
                row_start += 1				
                continue	

            lines_list.append((pid[0],qty))
			
            row_start += 1
			
        if error_list: 
            raise osv.except_osv( u'下列产品名错误', ', '.join(error_list) )	

        inventory_obj = self.pool.get('stock.inventory').browse(cr,uid,ids[0],context=context)
        inventory_line_obj = self.pool.get('stock.inventory.line')
        line_val = {
            'inventory_id':ids[0],
            'location_id': inventory_obj.location_id.id, 		
            'product_id': 0, 
            'product_uom_id': 1, 
            'product_qty': 0, 			
        }
		
        for line in lines_list :
            product = product_obj.browse(cr,uid,line[0],context=context)	
            if product.type != 'product' : continue			
            line_val.update({'product_id': line[0], 'total_all': line[1], 'product_qty': line[1] - product.qty_onhand_ec})		
            inventory_line_obj.create(cr,uid,line_val,context=context)
			
        attach.unlink()		
        return True					
			